import logging
import mimetypes
from io import BytesIO
from django import forms
from datetime import datetime, date as date_type, time as time_type
from django.conf import settings
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, Permission, User
from django.db import transaction
from django.db.models import Q, Count, Avg, Max, Prefetch
from django.db.models.functions import TruncMonth
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.cache import never_cache
from django.core.cache import cache
from django.views.decorators.csrf import ensure_csrf_cookie

from ..models import (
    Ciudadano, Atencion, Satisfaccion, Servicio, PrestamoRecurso,
    Recurso, PuntoViveDigital, Sala, UserProfile, AuditoriaAccion,
    PermisoDefinicion, PermisoRol, PermisoUsuario, HabilitacionSala,
    Curso, SesionCurso, InscripcionCurso, AsistenciaSesion,
    MantenimientoEquipo, Evidencia,
)
from ..forms import (
    CiudadanoForm, AtencionForm, SatisfaccionForm, ServicioForm,
    PrestamoRecursoForm, RecursoForm, LoginForm, PerfilUsuarioForm,
    CrearUsuarioForm, CrearUsuarioSistemaForm, EditarAdminPvdForm, PuntoViveDigitalForm, SalaForm, PermisoDefinicionForm,
    HabilitacionSalaForm, CursoForm, SesionCursoForm, InscripcionCursoForm,
    MantenimientoEquipoForm, EvidenciaForm,
)
from ..utils import registrar_auditoria, tiene_permiso, sincronizar_admin_a_cargo

logger = logging.getLogger('modulo_puntos')


from ._helpers import (
    usuario_es_superusuario, usuario_es_admin_tic,
    usuario_necesita_seleccionar_pvd, usuario_puede_usar_modulos_pvd,
    obtener_rol_usuario, obtener_pvd_activo_id, pvd_permitido,
    exigir_pvd_activo_para_admin_pvd,
)

# ── EXPORTACIÓN XLSX ───────────────────────────────────────────────────────────

def _xlsx_response(filename_base, wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}_{fecha_actual}.xlsx"'
    return response


def _estilizar_hoja(ws, headers, filas):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    cab_fondo = PatternFill(start_color='0B1220', end_color='0B1220', fill_type='solid')
    cab_fuente = Font(bold=True, color='FFFFFF', size=10)
    par_fondo = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    borde_cab = Border(
        left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='medium', color='4A90D9'),
    )
    borde_dato = Border(
        left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'),
    )
    for col, texto in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=texto)
        c.font = cab_fuente; c.fill = cab_fondo; c.border = borde_cab
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32
    for ri, fila in enumerate(filas, 2):
        for ci, valor in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=valor)
            c.alignment = Alignment(vertical='center', wrap_text=False)
            c.border = borde_dato
            if ri % 2 == 0:
                c.fill = par_fondo
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _crear_hoja(titulo, headers, filas):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = titulo[:31]
    _estilizar_hoja(wb.active, headers, filas)
    return wb


def _agregar_hoja(wb, titulo, headers, filas):
    ws = wb.create_sheet(titulo[:31])
    _estilizar_hoja(ws, headers, filas)
    return ws


@login_required(login_url='/login/')
def exportar_atenciones_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID Atención', 'Punto Vive Digital', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Estado',
        'Documento Ciudadano', 'Nombre Completo', 'Género', 'Etnia',
        'Discapacidad', 'Detalle Discapacidad', 'Barrio', 'Dirección',
        'Vereda / Corregimiento', 'Admin PVD a Cargo', 'Observaciones',
    ]
    estado_dict = dict(Atencion.ESTADO_CHOICES)
    qs = Atencion.objects.select_related('ciudadano', 'operador', 'punto_vive_digital').order_by('-fecha', '-hora_inicio')
    if pvd_id:
        qs = qs.filter(punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    filas = []
    for a in qs:
        c = a.ciudadano
        if c:
            doc = c.numero_documento or ''
            nom = f"{c.primer_nombre or ''} {c.primer_apellido or ''}".strip()
            gen = c.get_genero_display()
            etnia = c.etnia or ''
            discap = 'Sí' if c.tiene_discapacidad else 'No'
            desc_discap = c.descripcion_discapacidad or ''
            barrio = c.barrio or ''
            direccion = c.direccion or ''
            rural = c.zona_rural or ''
        else:
            doc = nom = gen = etnia = discap = desc_discap = barrio = direccion = rural = ''
        operador = a.operador.get_full_name() or a.operador.username if a.operador else ''
        pvd_nombre = a.punto_vive_digital.nombre if a.punto_vive_digital else ''
        filas.append([
            a.pk,
            pvd_nombre,
            str(a.fecha),
            str(a.hora_inicio),
            str(a.hora_fin) if a.hora_fin else '',
            estado_dict.get(a.estado, a.estado),
            doc, nom, gen, etnia, discap, desc_discap,
            barrio, direccion, rural, operador,
            a.observaciones or '',
        ])

    wb = _crear_hoja('Atenciones', headers, filas)
    return _xlsx_response('Reporte_Atenciones_PVD', wb)


@login_required(login_url='/login/')
def exportar_ciudadanos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    headers = [
        'ID', 'Punto Vive Digital', 'Tipo Documento', 'Número Documento',
        'Primer Nombre', 'Segundo Nombre', 'Primer Apellido', 'Segundo Apellido',
        'Fecha Nacimiento', 'Género', 'Etnia', 'Nivel Educativo', 'Ocupación',
        'Discapacidad', 'Descripción Discapacidad',
        'Dirección', 'Barrio', 'Zona Rural', 'Estrato', 'Estado', 'Email', 'Teléfono',
    ]
    qs = Ciudadano.objects.select_related('punto_vive_digital').order_by('-pk')
    if pvd_id:
        qs = qs.filter(punto_vive_digital_id=pvd_id)
    filas = []
    for c in qs:
        filas.append([
            c.pk,
            c.punto_vive_digital.nombre if c.punto_vive_digital else '',
            c.tipo_documento or '',
            c.numero_documento or '',
            c.primer_nombre or '',
            c.segundo_nombre or '',
            c.primer_apellido or '',
            c.segundo_apellido or '',
            str(c.fecha_nacimiento) if c.fecha_nacimiento else '',
            c.get_genero_display(),
            c.etnia or '',
            c.nivel_educativo or '',
            c.ocupacion or '',
            'Sí' if c.tiene_discapacidad else 'No',
            c.descripcion_discapacidad or '',
            c.direccion or '',
            c.barrio or '',
            c.zona_rural or '',
            c.estrato,
            c.get_estado_display(),
            c.correo or '',
            c.telefono or '',
        ])

    wb = _crear_hoja('Ciudadanos', headers, filas)
    return _xlsx_response('Reporte_Ciudadanos_PVD', wb)


@login_required(login_url='/login/')
def exportar_servicios_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID Servicio', 'Punto Vive Digital', 'ID Atención', 'Fecha Atención',
        'Documento Ciudadano', 'Nombre Ciudadano',
        'Nombre Servicio', 'Descripción', 'Tipo Servicio', 'Requiere Equipo', 'Recurso Utilizado', 'Estado',
    ]
    qs = Servicio.objects.select_related(
        'atencion', 'atencion__ciudadano', 'atencion__punto_vive_digital', 'recurso'
    ).order_by('-pk')
    if pvd_id:
        qs = qs.filter(atencion__punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs = qs.filter(atencion__fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(atencion__fecha__lte=fecha_hasta)
    filas = []
    for s in qs:
        atn = s.atencion
        fecha_atn = str(atn.fecha) if atn else ''
        pvd_nombre = atn.punto_vive_digital.nombre if atn and atn.punto_vive_digital else ''
        doc = nom = ''
        if atn and atn.ciudadano:
            doc = atn.ciudadano.numero_documento or ''
            nom = f"{atn.ciudadano.primer_nombre or ''} {atn.ciudadano.primer_apellido or ''}".strip()
        filas.append([
            s.pk,
            pvd_nombre,
            atn.pk if atn else '',
            fecha_atn,
            doc, nom,
            s.nombre,
            s.descripcion or '',
            s.tipo,
            s.get_requiere_equipo_display(),
            str(s.recurso) if s.recurso else '',
            s.get_estado_display(),
        ])

    wb = _crear_hoja('Servicios', headers, filas)
    return _xlsx_response('Reporte_Servicios_PVD', wb)


@login_required(login_url='/login/')
def exportar_satisfaccion_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    headers = [
        'ID', 'Punto Vive Digital', 'ID Atención', 'Fecha Atención', 'Estado Atención',
        'Documento Ciudadano', 'Nombre Ciudadano',
        '¿Tiempo de espera para ser atendido?', '¿Atención brindada por el servidor público?',
        '¿Quedó satisfecho con la prestación del servicio?', '¿La información recibida fue?',
        '¿Comodidad y limpieza de las instalaciones?', 'Puntaje promedio (1-5)',
        'Comentario', 'Fecha de Registro',
    ]
    estado_atn = dict(Atencion.ESTADO_CHOICES)
    qs = Satisfaccion.objects.select_related('atencion', 'atencion__ciudadano', 'atencion__punto_vive_digital').order_by('-pk')
    if pvd_id:
        qs = qs.filter(atencion__punto_vive_digital_id=pvd_id)
    filas = []
    for sat in qs:
        atn = sat.atencion
        doc = nom = fecha_atn = est_atn = pvd_nombre = ''
        if atn:
            fecha_atn = str(atn.fecha)
            est_atn = estado_atn.get(atn.estado, atn.estado)
            pvd_nombre = atn.punto_vive_digital.nombre if atn.punto_vive_digital else ''
            if atn.ciudadano:
                doc = atn.ciudadano.numero_documento or ''
                nom = f"{atn.ciudadano.primer_nombre or ''} {atn.ciudadano.primer_apellido or ''}".strip()
        promedio = sat.puntaje_promedio
        filas.append([
            sat.pk,
            pvd_nombre,
            atn.pk if atn else '',
            fecha_atn, est_atn,
            doc, nom,
            sat.get_tiempo_espera_display(),
            sat.get_atencion_servidor_display(),
            sat.get_satisfaccion_servicio_display(),
            sat.get_informacion_recibida_display(),
            sat.get_comodidad_instalaciones_display(),
            round(promedio, 1) if promedio is not None else '',
            sat.comentario or '',
            str(sat.fecha),
        ])

    wb = _crear_hoja('Satisfaccion', headers, filas)
    return _xlsx_response('Reporte_Satisfaccion_PVD', wb)


@login_required(login_url='/login/')
def exportar_prestamos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID Préstamo', 'Punto Vive Digital', 'Tipo Recurso', 'Código Recurso',
        'Fecha de Entrega', 'Fecha de Devolución', 'Observaciones', 'Estado',
    ]
    qs = PrestamoRecurso.objects.select_related('recurso', 'recurso__punto_vive_digital').order_by('-pk')
    if pvd_id:
        qs = qs.filter(recurso__punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs = qs.filter(fecha_entrega__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_entrega__lte=fecha_hasta)
    filas = []
    for p in qs:
        tipo = p.recurso.tipo if p.recurso else ''
        codigo = p.recurso.codigo or '' if p.recurso else ''
        pvd_nombre = p.recurso.punto_vive_digital.nombre if p.recurso and p.recurso.punto_vive_digital else ''
        estado = 'Activo (sin devolución)' if not p.fecha_devolucion else 'Devuelto'
        filas.append([
            p.pk,
            pvd_nombre,
            tipo,
            codigo,
            str(p.fecha_entrega),
            str(p.fecha_devolucion) if p.fecha_devolucion else '',
            p.observaciones or '',
            estado,
        ])

    wb = _crear_hoja('Prestamos', headers, filas)
    return _xlsx_response('Reporte_Prestamos_PVD', wb)


@login_required(login_url='/login/')
def exportar_cursos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    # Hoja 1 – Cursos / Talleres
    qs_cursos = (
        Curso.objects
        .select_related('punto_vive_digital', 'registrado_por')
        .order_by('-fecha_inicio')
    )
    if pvd_id:
        qs_cursos = qs_cursos.filter(punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs_cursos = qs_cursos.filter(fecha_inicio__gte=fecha_desde)
    if fecha_hasta:
        qs_cursos = qs_cursos.filter(fecha_inicio__lte=fecha_hasta)

    headers_cursos = [
        'ID', 'Punto Vive Digital', 'Nombre del Curso/Taller',
        'Modalidad', 'Población Objetivo', 'Fecha Inicio', 'Fecha Fin',
        'Estado', 'Total Inscritos', 'Registrado por',
    ]
    filas_cursos = []
    for c in qs_cursos:
        filas_cursos.append([
            c.pk,
            c.punto_vive_digital.nombre if c.punto_vive_digital else '',
            c.nombre,
            c.get_modalidad_display(),
            c.poblacion_objetivo or '',
            str(c.fecha_inicio),
            str(c.fecha_fin) if c.fecha_fin else '',
            c.get_estado_display(),
            c.total_inscritos(),
            c.registrado_por.get_full_name() or c.registrado_por.username if c.registrado_por else '',
        ])

    # Hoja 2 – Inscripciones
    qs_inscripciones = (
        InscripcionCurso.objects
        .select_related('curso', 'curso__punto_vive_digital', 'ciudadano', 'registrado_por')
        .order_by('curso__nombre', 'ciudadano__primer_apellido')
    )
    if pvd_id:
        qs_inscripciones = qs_inscripciones.filter(curso__punto_vive_digital_id=pvd_id)

    headers_inscripciones = [
        'ID Inscripción', 'Punto Vive Digital', 'Curso / Taller',
        'Documento Ciudadano', 'Nombre Ciudadano',
        'Estado Inscripción', 'Fecha Inscripción',
    ]
    filas_inscripciones = []
    for i in qs_inscripciones:
        ciu = i.ciudadano
        filas_inscripciones.append([
            i.pk,
            i.curso.punto_vive_digital.nombre if i.curso.punto_vive_digital else '',
            i.curso.nombre,
            ciu.numero_documento or '' if ciu else '',
            f"{ciu.primer_nombre or ''} {ciu.primer_apellido or ''}".strip() if ciu else '',
            i.get_estado_display(),
            str(i.fecha_inscripcion.date()) if i.fecha_inscripcion else '',
        ])

    # Hoja 3 – Asistencias por sesión
    qs_asistencias = (
        AsistenciaSesion.objects
        .select_related(
            'sesion', 'sesion__curso', 'sesion__curso__punto_vive_digital',
            'ciudadano',
        )
        .order_by('sesion__curso__nombre', 'sesion__numero_sesion', 'ciudadano__primer_apellido')
    )
    if pvd_id:
        qs_asistencias = qs_asistencias.filter(sesion__curso__punto_vive_digital_id=pvd_id)

    headers_asistencias = [
        'Punto Vive Digital', 'Curso / Taller', 'N° Sesión',
        'Fecha Sesión', 'Tema', 'Documento Ciudadano', 'Nombre Ciudadano', 'Asistió',
    ]
    filas_asistencias = []
    for a in qs_asistencias:
        ciu = a.ciudadano
        sesion = a.sesion
        filas_asistencias.append([
            sesion.curso.punto_vive_digital.nombre if sesion.curso.punto_vive_digital else '',
            sesion.curso.nombre,
            sesion.numero_sesion,
            str(sesion.fecha),
            sesion.tema,
            ciu.numero_documento or '' if ciu else '',
            f"{ciu.primer_nombre or ''} {ciu.primer_apellido or ''}".strip() if ciu else '',
            'Sí' if a.asistio else 'No',
        ])

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _agregar_hoja(wb, 'Cursos y Talleres', headers_cursos, filas_cursos)
    _agregar_hoja(wb, 'Inscripciones', headers_inscripciones, filas_inscripciones)
    _agregar_hoja(wb, 'Asistencias', headers_asistencias, filas_asistencias)
    return _xlsx_response('Reporte_Cursos_PVD', wb)


@login_required(login_url='/login/')
def exportar_mantenimientos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde_m = request.GET.get('fecha_desde', '').strip()
    fecha_hasta_m = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID', 'Punto Vive Digital', 'Tipo', 'Fecha',
        'Equipos Intervenidos', 'Descripción del Trabajo',
        'Hallazgos', 'Acciones / Recomendaciones', 'Registrado por',
    ]
    qs = (
        MantenimientoEquipo.objects
        .select_related('punto_vive_digital', 'realizado_por')
        .order_by('-fecha')
    )
    if pvd_id:
        qs = qs.filter(punto_vive_digital_id=pvd_id)
    if fecha_desde_m:
        qs = qs.filter(fecha__gte=fecha_desde_m)
    if fecha_hasta_m:
        qs = qs.filter(fecha__lte=fecha_hasta_m)

    filas = []
    for m in qs:
        filas.append([
            m.pk,
            m.punto_vive_digital.nombre if m.punto_vive_digital else '',
            m.get_tipo_display(),
            str(m.fecha),
            m.equipos_intervenidos,
            m.descripcion,
            m.hallazgos or '',
            m.acciones or '',
            m.realizado_por.get_full_name() or m.realizado_por.username if m.realizado_por else '',
        ])

    wb = _crear_hoja('Mantenimientos', headers, filas)
    return _xlsx_response('Reporte_Mantenimientos_PVD', wb)


