"""
Importa ciudadanos desde la plantilla oficial de Excel
(docs/Plantilla_Registro_Ciudadanos_PVD.xlsx) que llenan las
administradoras de cada PVD.

USO:
  # 1) Simulación (NO guarda nada — muestra qué pasaría):
  python manage.py importar_ciudadanos archivo.xlsx

  # 2) Importación real (solo cuando la simulación salga limpia):
  python manage.py importar_ciudadanos archivo.xlsx --confirmar

Cada fila se valida con las MISMAS reglas del formulario web
(CiudadanoForm): documento único, teléfono de 10 dígitos, tipo de
documento acorde a la edad, etc. Si una fila falla, se reporta con su
número de fila de Excel y no se importa; las demás sí.
"""
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from ...forms import CiudadanoForm
from ...models import Ciudadano, PuntoViveDigital, AuditoriaAccion

# Columnas de la hoja "Ciudadanos" (1-indexadas, igual que la plantilla)
COL = {
    'pvd': 1, 'tipo_doc': 2, 'numero_doc': 3, 'primer_nombre': 4,
    'segundo_nombre': 5, 'primer_apellido': 6, 'segundo_apellido': 7,
    'fecha_nac': 8, 'genero': 9, 'correo': 10, 'telefono': 11,
    'barrio': 12, 'municipio': 13, 'zona_rural': 14,
    'dir_tipo': 15, 'dir_num1': 16, 'dir_letra1': 17, 'dir_num2': 18,
    'dir_letra2': 19, 'dir_num3': 20, 'dir_comp': 21,
    # 22 = dirección armada (se recalcula aquí, no se confía en Excel)
    'etnia': 23, 'nivel_educativo': 24, 'ocupacion': 25, 'estrato': 26,
    'tiene_discapacidad': 27, 'desc_discapacidad': 28,
    'autorizacion': 29, 'estado': 30,
}

GENERO_MAP = {'Masculino': 'M', 'Femenino': 'F', 'Otro / No binario': 'O'}
ESTADO_MAP = {'Activo': 'A', 'Inactivo': 'I', '': 'A', None: 'A'}
FILA_INICIO = 4  # filas 1-3: encabezado, obligatoriedad y ejemplo

OBLIGATORIOS = [
    ('pvd', 'Punto Vive Digital'), ('tipo_doc', 'Tipo de Documento'),
    ('numero_doc', 'Número de Documento'), ('primer_nombre', 'Primer Nombre'),
    ('primer_apellido', 'Primer Apellido'), ('fecha_nac', 'Fecha de Nacimiento'),
    ('genero', 'Género'), ('barrio', 'Barrio'), ('etnia', 'Pertenencia Étnica'),
    ('nivel_educativo', 'Nivel Educativo'), ('ocupacion', 'Ocupación'),
    ('estrato', 'Estrato'), ('autorizacion', 'Autorización de datos'),
]


def _texto(valor):
    """Celda → texto limpio. Números enteros (documento, teléfono) sin '.0'."""
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    txt = _texto(valor)
    if txt:
        try:
            return datetime.strptime(txt, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def _extraer_codigo_doc(texto_doc):
    """'Cédula de Ciudadanía (CC)' → 'CC' (el código que guarda el sistema)."""
    if '(' in texto_doc and texto_doc.rstrip().endswith(')'):
        return texto_doc.rstrip()[texto_doc.rindex('(') + 1:-1].strip()
    return texto_doc  # ya viene como código


def _armar_direccion(fila):
    """Réplica exacta del constructor de direcciones de la pantalla de registro."""
    tipo = _texto(fila['dir_tipo'])
    num1 = _texto(fila['dir_num1'])
    letra1 = _texto(fila['dir_letra1'])
    num2 = _texto(fila['dir_num2'])
    letra2 = _texto(fila['dir_letra2'])
    num3 = _texto(fila['dir_num3'])
    comp = _texto(fila['dir_comp'])

    if not num1:
        return ''
    partes = f"{tipo} {num1} {letra1}"
    if num2:
        partes += f" # {num2} {letra2}"
    if num3:
        partes += f" - {num3}"
    if comp:
        partes += f" {comp}"
    return ' '.join(partes.split())


class Command(BaseCommand):
    help = 'Importa ciudadanos desde la plantilla oficial de Excel (simulación por defecto; use --confirmar para guardar).'

    def add_arguments(self, parser):
        parser.add_argument('archivo', help='Ruta del archivo .xlsx llenado con la plantilla oficial')
        parser.add_argument('--confirmar', action='store_true',
                            help='Guarda de verdad. Sin esta opción solo se simula y se muestra el reporte.')

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(opts['archivo'], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontró el archivo: {opts['archivo']}")
        except Exception as e:
            raise CommandError(f'No se pudo abrir el archivo (¿es un .xlsx válido?): {e}')

        if 'Ciudadanos' not in wb.sheetnames:
            raise CommandError('El archivo no tiene la hoja "Ciudadanos". Use la plantilla oficial.')
        ws = wb['Ciudadanos']

        pvds = {p.nombre.strip().upper(): p for p in PuntoViveDigital.objects.filter(estado='A')}

        validos, errores = [], []
        docs_en_archivo = set()

        for r in range(FILA_INICIO, ws.max_row + 1):
            fila = {k: ws.cell(row=r, column=c).value for k, c in COL.items()}

            # Fila totalmente vacía → se ignora en silencio
            if all(_texto(v) == '' for v in fila.values()):
                continue

            problemas = []

            faltan = [nombre for clave, nombre in OBLIGATORIOS if _texto(fila[clave]) == '']
            if faltan:
                problemas.append(f"faltan campos obligatorios: {', '.join(faltan)}")

            pvd = pvds.get(_texto(fila['pvd']).upper())
            if _texto(fila['pvd']) and not pvd:
                problemas.append(f"el PVD '{_texto(fila['pvd'])}' no existe o no está activo en el sistema")

            if _texto(fila['autorizacion']) and _texto(fila['autorizacion']) != 'Sí':
                problemas.append('sin autorización de datos (Ley 1581 de 2012) la persona no se puede registrar')

            genero = GENERO_MAP.get(_texto(fila['genero']))
            if _texto(fila['genero']) and not genero:
                problemas.append(f"género no reconocido: '{_texto(fila['genero'])}' (use la lista de la plantilla)")

            fecha_nac = _fecha(fila['fecha_nac'])
            if _texto(fila['fecha_nac']) and not fecha_nac:
                problemas.append('fecha de nacimiento inválida (use AAAA-MM-DD, ej: 1990-05-23)')

            numero_doc = _texto(fila['numero_doc'])
            if numero_doc and not numero_doc.isdigit():
                problemas.append('el número de documento debe tener solo dígitos, sin puntos')
            if numero_doc in docs_en_archivo:
                problemas.append(f'documento {numero_doc} repetido dentro del archivo')

            estrato_txt = _texto(fila['estrato'])
            try:
                estrato = int(float(estrato_txt)) if estrato_txt else None
            except ValueError:
                estrato = None
                problemas.append(f"estrato inválido: '{estrato_txt}' (debe ser 1 a 6)")

            if problemas:
                errores.append((r, '; '.join(problemas)))
                continue

            datos = {
                'tipo_documento': _extraer_codigo_doc(_texto(fila['tipo_doc'])),
                'numero_documento': numero_doc,
                'primer_nombre': _texto(fila['primer_nombre']),
                'segundo_nombre': _texto(fila['segundo_nombre']),
                'primer_apellido': _texto(fila['primer_apellido']),
                'segundo_apellido': _texto(fila['segundo_apellido']),
                'fecha_nacimiento': fecha_nac.isoformat(),
                'genero': genero,
                'correo': _texto(fila['correo']),
                'telefono': _texto(fila['telefono']),
                'direccion': _armar_direccion(fila),
                'municipio': _texto(fila['municipio']) or 'Bugalagrande',
                'barrio': _texto(fila['barrio']),
                'zona_rural': _texto(fila['zona_rural']),
                'etnia': _texto(fila['etnia']),
                'nivel_educativo': _texto(fila['nivel_educativo']),
                'ocupacion': _texto(fila['ocupacion']),
                'estrato': estrato,
                'tiene_discapacidad': _texto(fila['tiene_discapacidad']) == 'Sí',
                'descripcion_discapacidad': _texto(fila['desc_discapacidad']),
                'estado': ESTADO_MAP.get(_texto(fila['estado']), 'A'),
                'autorizacion_datos': True,
            }

            # Validación final con las MISMAS reglas del formulario web
            form = CiudadanoForm(data=datos)
            if not form.is_valid():
                detalle = '; '.join(
                    f"{campo}: {' '.join(str(e) for e in errs)}"
                    for campo, errs in form.errors.items()
                )
                errores.append((r, detalle))
                continue

            docs_en_archivo.add(numero_doc)
            validos.append((r, form, pvd))

        # ── Reporte ────────────────────────────────────────────────────────────
        modo = 'IMPORTACIÓN REAL' if opts['confirmar'] else 'SIMULACIÓN (no se guardó nada)'
        self.stdout.write(self.style.MIGRATE_HEADING(f'\n=== {modo} — {opts["archivo"]} ==='))
        self.stdout.write(f'Filas correctas:   {len(validos)}')
        self.stdout.write(f'Filas con errores: {len(errores)}')
        for fila_num, msg in errores:
            self.stdout.write(self.style.ERROR(f'  Fila {fila_num}: {msg}'))

        if not opts['confirmar']:
            if errores:
                self.stdout.write(self.style.WARNING(
                    '\nCorrija las filas con error en el Excel y vuelva a simular. '
                    'Cuando salga limpio, ejecute de nuevo con --confirmar.'
                ))
            elif validos:
                self.stdout.write(self.style.SUCCESS(
                    '\nSimulación limpia. Ejecute de nuevo agregando --confirmar para guardar.'
                ))
            else:
                self.stdout.write(self.style.WARNING('\nEl archivo no tiene filas de datos.'))
            return

        if not validos:
            self.stdout.write(self.style.WARNING('Nada para importar.'))
            return

        with transaction.atomic():
            for _, form, pvd in validos:
                ciudadano = form.save(commit=False)
                ciudadano.punto_vive_digital = pvd
                ciudadano.save()
            AuditoriaAccion.objects.create(
                usuario='importar_ciudadanos (comando)',
                accion='CREATE',
                modelo_afectado='Ciudadano',
                descripcion=(
                    f'Importación masiva desde plantilla Excel "{opts["archivo"]}": '
                    f'{len(validos)} ciudadanos creados, {len(errores)} filas rechazadas.'
                ),
            )
        self.stdout.write(self.style.SUCCESS(
            f'\n{len(validos)} ciudadanos importados correctamente.'
            + (f' {len(errores)} filas quedaron pendientes de corrección.' if errores else '')
        ))
