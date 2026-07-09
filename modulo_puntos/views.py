import logging
import mimetypes
from io import BytesIO
from django import forms
from datetime import datetime, date as date_type, time as time_type
from django.conf import settings
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, Permission, User
from django.db import transaction
from django.db.models import Q, Count, Avg, Max, Prefetch
from django.db.models.functions import TruncMonth
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.cache import never_cache
from django.core.cache import cache
from django.views.decorators.csrf import ensure_csrf_cookie

from .models import (
    Ciudadano, Atencion, Satisfaccion, Servicio, PrestamoRecurso,
    Recurso, PuntoViveDigital, Sala, UserProfile, AuditoriaAccion,
    PermisoDefinicion, PermisoRol, PermisoUsuario, HabilitacionSala,
    Curso, SesionCurso, InscripcionCurso, AsistenciaSesion,
    MantenimientoEquipo, Evidencia,
)
from .forms import (
    CiudadanoForm, AtencionForm, SatisfaccionForm, ServicioForm,
    PrestamoRecursoForm, RecursoForm, LoginForm, PerfilUsuarioForm,
    CrearUsuarioForm, CrearUsuarioSistemaForm, EditarAdminPvdForm, PuntoViveDigitalForm, SalaForm, PermisoDefinicionForm,
    HabilitacionSalaForm, CursoForm, SesionCursoForm, InscripcionCursoForm,
    MantenimientoEquipoForm, EvidenciaForm,
)
from .utils import registrar_auditoria, tiene_permiso, sincronizar_admin_a_cargo

logger = logging.getLogger('modulo_puntos')

# ── HELPERS ────────────────────────────────────────────────────────────────────

def usuario_es_superusuario(user):
    return user.is_authenticated and user.is_superuser


def usuario_es_admin_tic(user):
    if not user.is_authenticated:
        return False
    return user.is_superuser or user.groups.filter(name='Administrador TIC').exists()


def usuario_necesita_seleccionar_pvd(user):
    """Admin PVD debe elegir un PVD antes de operar."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or usuario_es_admin_tic(user):
        return False
    return user.groups.filter(name='Administrador PVD').exists()


def usuario_puede_usar_modulos_pvd(user):
    if not user.is_authenticated:
        return False
    return (
        user.is_superuser
        or usuario_es_admin_tic(user)
        or user.groups.filter(name='Administrador PVD').exists()
    )


def obtener_rol_usuario(user):
    if not user.is_authenticated:
        return 'Sin sesión'
    if user.is_superuser:
        return 'Superusuario'
    grupos = list(user.groups.values_list('name', flat=True))
    return ', '.join(grupos) if grupos else 'Sin rol asignado'


def obtener_pvd_activo_id(request):
    """PVD activo guardado en sesión (int) o None si no hay ninguno."""
    pvd_id = request.session.get('pvd_activo_id')
    try:
        return int(pvd_id) if pvd_id else None
    except (TypeError, ValueError):
        return None


def pvd_permitido(request, punto_vive_digital_id):
    """
    Verifica si el usuario puede ver/editar/eliminar un objeto que pertenece a
    `punto_vive_digital_id` (puede ser None si el objeto no tiene PVD asignado).

    - Superusuario y Administrador TIC: siempre True (ven todos los PVD).
    - Administrador PVD: sólo True si coincide con su PVD activo en sesión.
    """
    if not usuario_necesita_seleccionar_pvd(request.user):
        return True
    pvd_activo_id = obtener_pvd_activo_id(request)
    return pvd_activo_id is not None and punto_vive_digital_id == pvd_activo_id


def exigir_pvd_activo_para_admin_pvd(request):
    """
    Para reportes/exportaciones: si el usuario es Admin PVD y no tiene un PVD
    activo válido en sesión, retorna None (sin permitir "ver todo" por defecto).
    Retorna el pvd_id (int) a filtrar, o False si el usuario no está restringido
    (Superusuario/Admin TIC → sin filtro, ven todos los PVD).
    """
    if not usuario_necesita_seleccionar_pvd(request.user):
        return False
    return obtener_pvd_activo_id(request)


# ── AUTENTICACIÓN ──────────────────────────────────────────────────────────────


_LOGIN_MAX_INTENTOS = 5
_LOGIN_BLOQUEO_SEGUNDOS = 300  # 5 minutos


@never_cache
@ensure_csrf_cookie
def login_usuario(request):
    if request.user.is_authenticated:
        return redirect('modulo_puntos:panel_control')

    raw_next = request.GET.get('next') or request.POST.get('next') or ''
    if raw_next and url_has_allowed_host_and_scheme(raw_next, allowed_hosts={request.get_host()}):
        next_url = raw_next
    else:
        next_url = reverse('modulo_puntos:panel_control')

    # Control de intentos fallidos por IP (última entrada XFF = puesta por proxy, no por cliente)
    _xff = request.META.get('HTTP_X_FORWARDED_FOR', '')
    ip = _xff.split(',')[-1].strip() if _xff else request.META.get('REMOTE_ADDR', 'unknown')
    cache_key = f'login_intentos_{ip}'
    bloqueo_key = f'login_bloqueo_{ip}'

    pvd_count = PuntoViveDigital.objects.filter(estado='A').count()

    bloqueado_hasta = cache.get(bloqueo_key)
    segundos_restantes = 0
    if bloqueado_hasta:
        import time
        segundos_restantes = max(0, int(bloqueado_hasta - time.time()))
        if segundos_restantes > 0:
            minutos = segundos_restantes // 60
            segundos = segundos_restantes % 60
            messages.error(request, f'Demasiados intentos fallidos. Espera {minutos}m {segundos}s antes de intentar de nuevo.')
            return render(request, 'registration/login.html', {
                'form': LoginForm(),
                'next': next_url,
                'bloqueado': True,
                'segundos_restantes': segundos_restantes,
                'pvd_count': pvd_count,
            })

    form = LoginForm(request=request, data=request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            user = form.get_user()
            cache.delete(cache_key)
            cache.delete(bloqueo_key)
            login(request, user)
            messages.success(request, 'Inicio de sesión correcto.')
            if usuario_necesita_seleccionar_pvd(user):
                try:
                    profile = user.pvd_profile
                    if profile.punto_asignado_id and not profile.pvd_temporal_id:
                        pvd = profile.punto_asignado
                        if pvd.estado == 'A':
                            request.session['pvd_activo_id'] = pvd.pk
                            request.session['pvd_nombre'] = pvd.nombre
                            messages.success(request, f'Trabajando en: {pvd.nombre}')
                            return redirect(next_url)
                except UserProfile.DoesNotExist:
                    pass
                return redirect('modulo_puntos:seleccionar_pvd_view')
            return redirect(next_url)
        else:
            import time
            intentos = cache.get(cache_key, 0) + 1
            cache.set(cache_key, intentos, _LOGIN_BLOQUEO_SEGUNDOS)
            restantes = _LOGIN_MAX_INTENTOS - intentos
            if intentos >= _LOGIN_MAX_INTENTOS:
                cache.set(bloqueo_key, time.time() + _LOGIN_BLOQUEO_SEGUNDOS, _LOGIN_BLOQUEO_SEGUNDOS)
                cache.delete(cache_key)
                messages.error(request, f'Cuenta bloqueada por {_LOGIN_BLOQUEO_SEGUNDOS // 60} minutos por múltiples intentos fallidos.')
            else:
                messages.error(request, f'Usuario o contraseña incorrectos. Te quedan {restantes} intento(s).')

    return render(request, 'registration/login.html', {'form': form, 'next': next_url, 'pvd_count': pvd_count})


@login_required(login_url='/login/')
def logout_usuario(request):
    if request.method != 'POST':
        return redirect('modulo_puntos:panel_control')
    logout(request)
    return redirect('modulo_puntos:login')


# ── PERFIL ─────────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def perfil_usuario(request):
    if request.method == 'POST':
        if request.POST.get('cambiar_password'):
            old_pw = request.POST.get('old_password', '').strip()
            new_pw1 = request.POST.get('new_password1', '').strip()
            new_pw2 = request.POST.get('new_password2', '').strip()
            if not request.user.check_password(old_pw):
                messages.error(request, 'La contraseña actual no es correcta.')
            elif len(new_pw1) < 8:
                messages.error(request, 'La nueva contraseña debe tener al menos 8 caracteres.')
            elif new_pw1 != new_pw2:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            else:
                request.user.set_password(new_pw1)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña cambiada correctamente.')
            return redirect('modulo_puntos:perfil_usuario')

        form = PerfilUsuarioForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil fue actualizado correctamente.')
            return redirect('modulo_puntos:perfil_usuario')
        messages.error(request, 'No se pudo actualizar el perfil. Revisa los datos ingresados.')
    else:
        form = PerfilUsuarioForm(instance=request.user)

    return render(request, 'modulo_puntos/perfil_usuario.html', {
        'form': form,
        'rol_usuario': obtener_rol_usuario(request.user),
    })


# ── PANEL DE CONTROL ───────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def panel_control(request):
    user = request.user

    # Admin PVD debe seleccionar PVD antes de entrar al panel
    if usuario_necesita_seleccionar_pvd(user):
        pvd_id_session = request.session.get('pvd_activo_id')
        if not pvd_id_session:
            messages.info(request, 'Selecciona un Punto Vive Digital para comenzar.')
            return redirect('modulo_puntos:seleccionar_pvd_view')

    from datetime import date as date_cls
    hoy = date_cls.today()
    context = {
        'total_ciudadanos': Ciudadano.objects.count(),
        'atenciones_registradas': Atencion.objects.count(),
        'total_satisfacciones': Satisfaccion.objects.count(),
        'total_servicios': Servicio.objects.count(),
        'total_prestamos': PrestamoRecurso.objects.count(),
        'total_pvds': PuntoViveDigital.objects.count(),
        'total_recursos': Recurso.objects.count(),
        'total_cursos': Curso.objects.count(),
        'atenciones_hoy_global': Atencion.objects.filter(fecha=hoy).count(),

        'es_superusuario': usuario_es_superusuario(user),
        'es_admin_tic_only': user.groups.filter(name='Administrador TIC').exists() and not user.is_superuser,
        'es_admin_pvd_only': user.groups.filter(name='Administrador PVD').exists() and not usuario_es_admin_tic(user),
        'mostrar_modulos_tic': usuario_es_admin_tic(user),
        'mostrar_modulos_pvd': usuario_puede_usar_modulos_pvd(user),
        'rol_usuario': obtener_rol_usuario(user),
        'pvd_activo': None,
    }
    pvd_id = request.session.get('pvd_activo_id')
    if pvd_id:
        pvd_activo = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first()
        context['pvd_activo'] = pvd_activo
        if pvd_activo:
            context['pvd_ciudadanos']     = Ciudadano.objects.filter(punto_vive_digital_id=pvd_id).count()
            context['pvd_atenciones']     = Atencion.objects.filter(punto_vive_digital_id=pvd_id).count()
            context['pvd_salas']          = Sala.objects.filter(punto_vive_digital_id=pvd_id).count()
            context['pvd_cursos']         = Curso.objects.filter(punto_vive_digital_id=pvd_id).count()
            context['pvd_mantenimientos'] = MantenimientoEquipo.objects.filter(punto_vive_digital_id=pvd_id).count()
            context['pvd_habilitaciones'] = HabilitacionSala.objects.filter(sala__punto_vive_digital_id=pvd_id).count()
            context['atenciones_hoy'] = Atencion.objects.filter(punto_vive_digital_id=pvd_id, fecha=hoy).count()
            context['atenciones_pendientes'] = (
                Atencion.objects
                .filter(punto_vive_digital_id=pvd_id, estado='P')
                .select_related('ciudadano')
                .order_by('-fecha', '-pk')[:10]
            )
    return render(request, 'modulo_puntos/panel_control.html', context)


# ── SELECCIÓN DE PVD ───────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def seleccionar_pvd_view(request):
    sin_asignacion = False
    permanente_id = None
    temporal_id = None

    conteos = dict(
        total_ciudadanos=Count('ciudadano', filter=Q(ciudadano__estado='A'), distinct=True),
        total_atenciones=Count('atencion', distinct=True),
        total_recursos=Count('recurso', filter=Q(recurso__estado='A'), distinct=True),
        total_salas=Count('sala', filter=Q(sala__estado='A'), distinct=True),
    )

    if usuario_necesita_seleccionar_pvd(request.user):
        try:
            profile = request.user.pvd_profile
            if profile.punto_asignado_id:
                permanente_id = profile.punto_asignado_id
                pvd_ids = [profile.punto_asignado_id]
                if profile.pvd_temporal_id:
                    temporal_id = profile.pvd_temporal_id
                    pvd_ids.append(profile.pvd_temporal_id)
                pvds = PuntoViveDigital.objects.filter(pk__in=pvd_ids, estado='A').annotate(**conteos).order_by('nombre')
            else:
                # Admin PVD sin PVD asignado: bloquear, no mostrar lista
                pvds = PuntoViveDigital.objects.none()
                sin_asignacion = True
        except UserProfile.DoesNotExist:
            pvds = PuntoViveDigital.objects.none()
            sin_asignacion = True
    else:
        pvds = PuntoViveDigital.objects.filter(estado='A').annotate(**conteos).order_by('nombre')

    pvds = list(pvds)
    for pvd in pvds:
        pvd.es_permanente = (pvd.pk == permanente_id)
        pvd.es_temporal = (pvd.pk == temporal_id)

    return render(request, 'modulo_puntos/seleccionar_pvd.html', {
        'pvds': pvds,
        'sin_asignacion': sin_asignacion,
    })


@login_required(login_url='/login/')
def seleccionar_pvd(request, pvd_cdgo):
    pvd = get_object_or_404(PuntoViveDigital, pk=pvd_cdgo, estado='A')

    if usuario_necesita_seleccionar_pvd(request.user):
        try:
            profile = request.user.pvd_profile
            if not profile.punto_asignado_id:
                messages.error(request, 'No tienes ningún Punto Vive Digital asignado. Contacta al Administrador TIC.')
                return redirect('modulo_puntos:seleccionar_pvd_view')
            pvd_ids_permitidos = [profile.punto_asignado_id]
            if profile.pvd_temporal_id:
                pvd_ids_permitidos.append(profile.pvd_temporal_id)
            if pvd_cdgo not in pvd_ids_permitidos:
                messages.error(request, 'No tienes permiso para acceder a ese Punto Vive Digital.')
                return redirect('modulo_puntos:seleccionar_pvd_view')
        except UserProfile.DoesNotExist:
            messages.error(request, 'No tienes ningún Punto Vive Digital asignado. Contacta al Administrador TIC.')
            return redirect('modulo_puntos:seleccionar_pvd_view')

    request.session['pvd_activo_id'] = pvd.pk
    request.session['pvd_nombre'] = pvd.nombre
    messages.success(request, f'Trabajando en: {pvd.nombre}')
    return redirect('modulo_puntos:panel_control')


# ── GESTIÓN DE CIUDADANOS ──────────────────────────────────────────────────────

@login_required(login_url='/login/')
def consultar_ciudadanos(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    busqueda   = request.GET.get('q', '').strip()
    pvd_filtro = request.GET.get('pvd', '').strip()
    estado_filtro = request.GET.get('estado', 'A').strip()  # Por defecto solo activos

    # Los ciudadanos son una población compartida: cualquier PVD puede consultar,
    # atender y editar a un ciudadano sin importar en qué PVD fue registrado.
    ciudadanos = Ciudadano.objects.select_related('punto_vive_digital').order_by('-pk')

    if pvd_filtro:
        ciudadanos = ciudadanos.filter(punto_vive_digital_id=pvd_filtro)

    if estado_filtro in ('A', 'I'):
        ciudadanos = ciudadanos.filter(estado=estado_filtro)
    # '' → todos (activos + inactivos)

    if busqueda:
        ciudadanos = ciudadanos.filter(
            Q(numero_documento__icontains=busqueda) |
            Q(primer_nombre__icontains=busqueda)   |
            Q(primer_apellido__icontains=busqueda)  |
            Q(segundo_nombre__icontains=busqueda)   |
            Q(segundo_apellido__icontains=busqueda) |
            Q(telefono__icontains=busqueda)         |
            Q(barrio__icontains=busqueda)
        )

    total_resultados = ciudadanos.count()
    paginator = Paginator(ciudadanos, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    pvds_para_filtro = PuntoViveDigital.objects.filter(estado='A').order_by('nombre')

    return render(request, 'modulo_puntos/consultar_ciudadanos.html', {
        'ciudadanos': page_obj,
        'page_obj': page_obj,
        'busqueda': busqueda,
        'pvd_filtro': pvd_filtro,
        'estado_filtro': estado_filtro,
        'pvds_para_filtro': pvds_para_filtro,
        'total_resultados': total_resultados,
    })


@login_required(login_url='/login/')
def registrar_ciudadano(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    # Restricción de sesión: sin PVD activo no se puede registrar ningún ciudadano,
    # independientemente del rol (incluyendo superusuario).
    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(
            request,
            'Debes ingresar a un Punto Vive Digital antes de registrar un ciudadano. '
            'Selecciona uno de la lista.'
        )
        return redirect('modulo_puntos:seleccionar_pvd_view')

    pvd_activo = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first()
    if not pvd_activo:
        # El PVD en sesión fue desactivado o eliminado
        del request.session['pvd_activo_id']
        messages.warning(request, 'El Punto Vive Digital de tu sesión ya no está disponible. Selecciona otro.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    form = CiudadanoForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            try:
                ciudadano = form.save(commit=False)
                ciudadano.punto_vive_digital_id = pvd_id   # siempre obligatorio
                ciudadano.save()
                registrar_auditoria(
                    request, 'CREATE', 'Ciudadano', ciudadano.pk,
                    f'Ciudadano registrado en {pvd_activo.nombre}: {ciudadano.get_nombre_completo()}'
                )
                messages.success(request, 'Ciudadano registrado exitosamente en la base de datos.')
                return redirect('modulo_puntos:panel_control')
            except Exception as e:
                logger.error('Error al guardar ciudadano: %s', e, exc_info=True)
                messages.error(request, 'Error al guardar los datos. Inténtalo de nuevo.')
        else:
            messages.error(request, 'Formulario inválido. Revisa los campos.')

    return render(request, 'modulo_puntos/registrar_ciudadano.html', {
        'form': form,
        'pvd_activo': pvd_activo,
    })


@login_required(login_url='/login/')
def editar_ciudadano(request, ciu_cdgo):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    # Los ciudadanos son una población compartida entre PVDs: cualquier Admin PVD
    # puede editar los datos de un ciudadano, sin importar dónde fue registrado.
    ciudadano = get_object_or_404(Ciudadano, pk=ciu_cdgo)

    form = CiudadanoForm(request.POST or None, instance=ciudadano)
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Ciudadano actualizado correctamente en la base de datos.')
                return redirect('modulo_puntos:consultar_ciudadanos')
            except Exception as e:
                logger.error('Error al actualizar ciudadano %s: %s', ciu_cdgo, e, exc_info=True)
                messages.error(request, 'Error al guardar los datos. Inténtalo de nuevo.')
        else:
            messages.error(request, 'No se pudo actualizar el ciudadano. Revisa los datos ingresados.')

    return render(request, 'modulo_puntos/editar_ciudadano.html', {
        'form': form,
        'ciudadano': ciudadano,
    })


@login_required(login_url='/login/')
def desactivar_ciudadano(request, ciu_cdgo):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if request.method != 'POST':
        return redirect('modulo_puntos:consultar_ciudadanos')
    # Los ciudadanos son una población compartida entre PVDs.
    ciudadano = get_object_or_404(Ciudadano, pk=ciu_cdgo)
    nuevo_estado = 'I' if ciudadano.estado == 'A' else 'A'
    ciudadano.estado = nuevo_estado
    ciudadano.save(update_fields=['estado'])
    etiqueta = 'desactivado' if nuevo_estado == 'I' else 'reactivado'
    registrar_auditoria(request, 'UPDATE', 'Ciudadano', ciudadano.pk,
                        f'Ciudadano {etiqueta}: {ciudadano.get_nombre_completo()}')
    messages.success(request, f'Ciudadano {etiqueta} correctamente.')
    return redirect('modulo_puntos:consultar_ciudadanos')


@login_required(login_url='/login/')
def historial_ciudadano(request, ciu_cdgo):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    # Los ciudadanos son una población compartida entre PVDs: cualquier Admin PVD
    # puede consultar el historial completo de un ciudadano (incluye atenciones
    # prestadas en cualquier PVD), sin importar dónde fue registrado.
    ciudadano = get_object_or_404(Ciudadano, pk=ciu_cdgo)

    atenciones = list(
        Atencion.objects.filter(ciudadano=ciudadano)
        .select_related('operador', 'prestamo', 'prestamo__recurso')
        .prefetch_related(
            Prefetch('servicio_set', queryset=Servicio.objects.order_by('pk'), to_attr='servicios_rel'),
            Prefetch('satisfaccion_set', to_attr='satisfacciones_rel'),
        )
        .order_by('-fecha', '-hora_inicio')
    )

    return render(request, 'modulo_puntos/historial_ciudadano.html', {
        'ciudadano': ciudadano,
        'atenciones': atenciones,
        'total_atenciones': len(atenciones),
    })


@login_required(login_url='/login/')
def ciudadanos_pendientes(request):
    if not (request.user.is_superuser or usuario_es_admin_tic(request.user)):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pendientes = list(Ciudadano.objects.filter(estado='P').order_by('-fecha_registro'))
    return render(request, 'modulo_puntos/ciudadanos_pendientes.html', {
        'ciudadanos_pendientes': pendientes,
        'total_pendientes': len(pendientes),
    })


@login_required(login_url='/login/')
def aprobar_ciudadano(request, ciu_id):
    if not (request.user.is_superuser or usuario_es_admin_tic(request.user)):
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('modulo_puntos:panel_control')

    if request.method != 'POST':
        return redirect('modulo_puntos:ciudadanos_pendientes')

    ciudadano = get_object_or_404(Ciudadano, pk=ciu_id, estado='P')
    nombre = ciudadano.get_nombre_completo()
    ciudadano.estado = 'A'
    ciudadano.save()
    messages.success(request, f'Ciudadano {nombre} aprobado exitosamente.')
    return redirect('modulo_puntos:ciudadanos_pendientes')


@login_required(login_url='/login/')
def rechazar_ciudadano(request, ciu_id):
    if not (request.user.is_superuser or usuario_es_admin_tic(request.user)):
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('modulo_puntos:panel_control')

    if request.method != 'POST':
        return redirect('modulo_puntos:ciudadanos_pendientes')

    ciudadano = get_object_or_404(Ciudadano, pk=ciu_id, estado='P')
    nombre = ciudadano.get_nombre_completo()
    ciudadano.delete()
    messages.success(request, f'Registro de {nombre} rechazado y eliminado.')
    return redirect('modulo_puntos:ciudadanos_pendientes')


# ── REGISTRO CIUDADANO PÚBLICO (sin login) ─────────────────────────────────────

def registrar_usuario_ciudadano(request):
    _xff = request.META.get('HTTP_X_FORWARDED_FOR', '')
    _ip = _xff.split(',')[-1].strip() if _xff else request.META.get('REMOTE_ADDR', 'unknown')
    _rate_key = f'reg_pub_{_ip}'
    if cache.get(_rate_key, 0) >= 5:
        return render(request, 'modulo_puntos/registrar_usuario_ciudadano.html', {
            'form': CiudadanoForm(),
            'bloqueado': True,
        }, status=429)

    form = CiudadanoForm(request.POST or None)
    if request.method == 'POST':
        cache.set(_rate_key, cache.get(_rate_key, 0) + 1, 300)
        if form.is_valid():
            ciudadano = form.save(commit=False)
            ciudadano.estado = 'P'
            # Autorregistro público: no hay PVD activo en sesión (no requiere login),
            # así que se asigna al primer PVD activo por defecto; el Admin PVD
            # puede reasignarlo al aprobar el registro.
            ciudadano.punto_vive_digital = PuntoViveDigital.objects.filter(estado='A').order_by('nombre').first()
            ciudadano.save()
            return redirect('modulo_puntos:registro_exitoso')
        messages.error(request, 'Revisa los datos ingresados.')
    return render(request, 'modulo_puntos/registrar_usuario_ciudadano.html', {'form': form})


def registro_exitoso(request):
    return render(request, 'modulo_puntos/registro_exitoso.html')


# ── REPORTES ───────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def reportes(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    fecha_desde = None
    fecha_hasta = None
    try:
        if fecha_desde_str:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
        if fecha_hasta_str:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
    except ValueError:
        pass

    # Filtrar por PVD activo cuando corresponde
    es_admin_pvd_rpt = usuario_necesita_seleccionar_pvd(request.user)
    pvd_id_rpt = obtener_pvd_activo_id(request)
    if es_admin_pvd_rpt and not pvd_id_rpt:
        # Fail-closed: sin PVD activo en sesión, un Admin PVD no puede ver reportes
        # (antes esto dejaba los querysets sin filtrar y exponía todos los PVD).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de ver los reportes.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    pvd_nombre_rpt = None
    if pvd_id_rpt:
        pvd_obj = PuntoViveDigital.objects.filter(pk=pvd_id_rpt).first()
        pvd_nombre_rpt = pvd_obj.nombre if pvd_obj else None

    atencion_qs = Atencion.objects.all()
    servicio_qs = Servicio.objects.all()
    ciudadano_qs = Ciudadano.objects.all()
    prestamo_qs  = PrestamoRecurso.objects.all()
    satisfaccion_qs = Satisfaccion.objects.all()

    if pvd_id_rpt:
        # Admin PVD siempre filtrado; Super/TIC filtrados sólo si eligieron un PVD activo.
        atencion_qs  = atencion_qs.filter(punto_vive_digital_id=pvd_id_rpt)
        servicio_qs  = servicio_qs.filter(atencion__punto_vive_digital_id=pvd_id_rpt)
        ciudadano_qs = ciudadano_qs.filter(punto_vive_digital_id=pvd_id_rpt)
        prestamo_qs  = prestamo_qs.filter(recurso__punto_vive_digital_id=pvd_id_rpt)
        satisfaccion_qs = satisfaccion_qs.filter(atencion__punto_vive_digital_id=pvd_id_rpt)

    if fecha_desde:
        atencion_qs = atencion_qs.filter(fecha__gte=fecha_desde)
        servicio_qs = servicio_qs.filter(atencion__fecha__gte=fecha_desde)
    if fecha_hasta:
        atencion_qs = atencion_qs.filter(fecha__lte=fecha_hasta)
        servicio_qs = servicio_qs.filter(atencion__fecha__lte=fecha_hasta)

    total_ciudadanos = ciudadano_qs.count()
    ciudadanos_activos = ciudadano_qs.filter(estado='A').count()
    total_atenciones = atencion_qs.count()
    atenciones_pendientes = atencion_qs.filter(estado='P').count()
    atenciones_finalizadas = atencion_qs.filter(estado='F').count()
    atenciones_canceladas = atencion_qs.filter(estado='C').count()
    total_servicios = servicio_qs.count()
    total_prestamos = prestamo_qs.count()
    prestamos_activos = prestamo_qs.filter(fecha_devolucion__isnull=True).count()

    satisfaccion_promedio = Satisfaccion.con_puntaje(satisfaccion_qs).aggregate(
        promedio=Avg('puntaje')
    )['promedio']

    servicios_por_tipo = servicio_qs.values('tipo').annotate(
        total=Count('id')
    ).order_by('-total', 'tipo')

    atenciones_por_admin = atencion_qs.values(
        'operador__first_name',
        'operador__last_name',
        'operador__username'
    ).annotate(
        total=Count('id')
    ).order_by('-total')

    atenciones_recientes = atencion_qs.select_related(
        'ciudadano', 'operador'
    ).order_by('-fecha', '-hora_inicio')[:10]

    gen_map = dict(Ciudadano.GENERO_CHOICES)
    ciudadanos_por_genero = []
    for row in ciudadano_qs.values('genero').annotate(total=Count('id')).order_by('-total'):
        clave = row['genero'] or ''
        ciudadanos_por_genero.append({
            'etiqueta': gen_map.get(clave, clave or 'Sin dato'),
            'total': row['total'],
        })

    ciudadanos_por_etnia = list(
        ciudadano_qs.exclude(etnia__isnull=True).exclude(etnia='').values('etnia').annotate(
            total=Count('id')
        ).order_by('-total')[:20]
    )

    ciudadanos_por_nvleduc = list(
        ciudadano_qs.exclude(nivel_educativo__isnull=True).exclude(nivel_educativo='').values('nivel_educativo').annotate(
            total=Count('id')
        ).order_by('-total')[:20]
    )

    ciudadanos_por_estrato = list(
        ciudadano_qs.values('estrato').annotate(total=Count('id')).order_by('estrato')
    )

    ciudadanos_por_ocupacion = list(
        ciudadano_qs.exclude(ocupacion__isnull=True).exclude(ocupacion='').values('ocupacion').annotate(
            total=Count('id')
        ).order_by('-total')[:15]
    )

    ciudadanos_con_discapacidad = ciudadano_qs.filter(tiene_discapacidad=True).count()
    ciudadanos_sin_discapacidad = ciudadano_qs.filter(tiene_discapacidad=False).count()

    atenciones_por_mes = list(
        atencion_qs.annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('-mes')[:12]
    )

    # Reporte de salas
    habilitacion_qs = HabilitacionSala.objects.all()
    if pvd_id_rpt:
        habilitacion_qs = habilitacion_qs.filter(sala__punto_vive_digital_id=pvd_id_rpt)
    if fecha_desde:
        habilitacion_qs = habilitacion_qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        habilitacion_qs = habilitacion_qs.filter(fecha__lte=fecha_hasta)

    uso_por_sala = list(
        habilitacion_qs
        .exclude(estado='X')
        .values('sala__nombre', 'sala__capacidad')
        .annotate(total_usos=Count('id'), total_personas=Avg('capacidad_requerida'))
        .order_by('-total_usos')[:10]
    )
    tipo_uso_map = dict(HabilitacionSala.TIPO_USO_CHOICES)
    uso_por_tipo = list(
        habilitacion_qs
        .exclude(estado='X')
        .values('tipo_uso')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    for item in uso_por_tipo:
        item['tipo_uso_label'] = tipo_uso_map.get(item['tipo_uso'], item['tipo_uso'])

    # KPIs operacionales
    from django.db.models import ExpressionWrapper, DurationField, F
    tasa_finalizacion = (
        round(atenciones_finalizadas * 100 / total_atenciones, 1)
        if total_atenciones > 0 else 0
    )
    tasa_cancelacion = (
        round(atenciones_canceladas * 100 / total_atenciones, 1)
        if total_atenciones > 0 else 0
    )

    satisfaccion_por_admin = list(
        Satisfaccion.con_puntaje(satisfaccion_qs)
        .filter(atencion__operador__isnull=False)
        .values('atencion__operador__first_name', 'atencion__operador__last_name', 'atencion__operador__username')
        .annotate(promedio=Avg('puntaje'), total=Count('id'))
        .order_by('-promedio')
    )

    return render(request, 'modulo_puntos/reportes.html', {
        'total_ciudadanos': total_ciudadanos,
        'ciudadanos_activos': ciudadanos_activos,
        'total_atenciones': total_atenciones,
        'atenciones_pendientes': atenciones_pendientes,
        'atenciones_finalizadas': atenciones_finalizadas,
        'atenciones_canceladas': atenciones_canceladas,
        'total_servicios': total_servicios,
        'total_prestamos': total_prestamos,
        'prestamos_activos': prestamos_activos,
        'satisfaccion_promedio': satisfaccion_promedio,
        'servicios_por_tipo': servicios_por_tipo,
        'atenciones_por_admin': atenciones_por_admin,
        'atenciones_recientes': atenciones_recientes,
        'ciudadanos_por_genero': ciudadanos_por_genero,
        'ciudadanos_por_etnia': ciudadanos_por_etnia,
        'ciudadanos_por_nvleduc': ciudadanos_por_nvleduc,
        'ciudadanos_por_estrato': ciudadanos_por_estrato,
        'ciudadanos_por_ocupacion': ciudadanos_por_ocupacion,
        'ciudadanos_con_discapacidad': ciudadanos_con_discapacidad,
        'ciudadanos_sin_discapacidad': ciudadanos_sin_discapacidad,
        'atenciones_por_mes': atenciones_por_mes,
        'fecha_desde': fecha_desde_str,
        'fecha_hasta': fecha_hasta_str,
        'pvd_nombre_rpt': pvd_nombre_rpt,
        'tasa_finalizacion': tasa_finalizacion,
        'tasa_cancelacion': tasa_cancelacion,
        'satisfaccion_por_admin': satisfaccion_por_admin,
        'uso_por_sala': uso_por_sala,
        'uso_por_tipo': uso_por_tipo,
        'tipo_uso_map': tipo_uso_map,
    })


# ── ATENCIONES Y SERVICIOS ─────────────────────────────────────────────────────

@login_required(login_url='/login/')
def buscar_ciudadanos_json(request):
    """Endpoint AJAX para el autocompletado de ciudadanos en el formulario de atención."""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    # Los ciudadanos son una población compartida: se pueden buscar y atender
    # desde cualquier PVD, sin importar en cuál fueron registrados originalmente.
    qs = Ciudadano.objects.filter(estado='A').filter(
        Q(primer_nombre__icontains=q)
        | Q(segundo_nombre__icontains=q)
        | Q(primer_apellido__icontains=q)
        | Q(segundo_apellido__icontains=q)
        | Q(numero_documento__icontains=q)
    ).order_by('primer_apellido', 'primer_nombre')[:12]

    return JsonResponse({'results': [
        {
            'id':     c.pk,
            'nombre': c.get_nombre_completo(),
            'doc':    c.numero_documento or '',
        }
        for c in qs
    ]})


@login_required(login_url='/login/')
def registrar_atencion(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(request, 'Debes ingresar a un Punto Vive Digital antes de registrar una atención.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    pvd_activo = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first()
    if not pvd_activo:
        del request.session['pvd_activo_id']
        messages.warning(request, 'El Punto Vive Digital de tu sesión ya no está disponible. Selecciona otro.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    ciudadano_id = request.GET.get('ciudadano')
    ciudadano_prefill = None
    initial = {}
    if ciudadano_id:
        try:
            ciudadano_prefill = Ciudadano.objects.get(pk=ciudadano_id, estado='A')
            initial['ciudadano'] = ciudadano_prefill
        except Ciudadano.DoesNotExist:
            pass

    form = AtencionForm(request.POST or None, initial=initial)

    if request.method == 'POST':
        if form.is_valid():
            try:
                atencion = form.save(commit=False)
                atencion.operador = request.user
                atencion.punto_vive_digital_id = pvd_id
                atencion.save()
                messages.success(request, 'Atención registrada. Ahora registra los servicios prestados.')
                return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
            except Exception as e:
                messages.error(request, f'Error al guardar en BD: {e}')
        else:
            messages.error(request, 'No se pudo guardar la atención. Revisa los datos ingresados.')

    return render(request, 'modulo_puntos/registrar_atencion.html', {
        'form': form,
        'ciudadano_prefill': ciudadano_prefill,
        'pvd_activo': pvd_activo,
    })


@login_required(login_url='/login/')
def editar_atencion(request, atencion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    atencion = get_object_or_404(Atencion, pk=atencion_id)

    # Admin PVD solo puede editar atenciones de su PVD activo
    pvd_id = request.session.get('pvd_activo_id')
    es_admin_pvd_solo = (
        not request.user.is_superuser
        and not usuario_es_admin_tic(request.user)
        and request.user.groups.filter(name='Administrador PVD').exists()
    )
    if es_admin_pvd_solo and pvd_id and atencion.punto_vive_digital_id != int(pvd_id):
        messages.error(request, 'No tienes permiso para editar esta atención.')
        return redirect('modulo_puntos:lista_atenciones')

    form = AtencionForm(request.POST or None, instance=atencion)

    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Atención actualizada correctamente.')
                return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
            except Exception as e:
                messages.error(request, f'Error al guardar: {e}')
        else:
            messages.error(request, 'Revisa los datos ingresados.')

    return render(request, 'modulo_puntos/registrar_atencion.html', {
        'form': form,
        'atencion': atencion,
    })


@login_required(login_url='/login/')
def lista_atenciones(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    user = request.user
    es_admin_pvd_solo = (
        not user.is_superuser
        and not usuario_es_admin_tic(user)
        and user.groups.filter(name='Administrador PVD').exists()
    )

    q = request.GET.get('q', '').strip()
    estado_filter = request.GET.get('estado', '').strip()
    fecha_desde_filter = request.GET.get('fecha_desde', '').strip()
    fecha_hasta_filter = request.GET.get('fecha_hasta', '').strip()
    pvd_filter = request.GET.get('pvd', '').strip()

    atenciones = Atencion.objects.select_related(
        'ciudadano', 'punto_vive_digital', 'operador'
    ).order_by('-fecha', '-pk')

    if es_admin_pvd_solo:
        pvd_id = request.session.get('pvd_activo_id')
        atenciones = atenciones.filter(punto_vive_digital_id=pvd_id) if pvd_id else atenciones.none()
    elif pvd_filter:
        atenciones = atenciones.filter(punto_vive_digital_id=pvd_filter)

    if q:
        atenciones = atenciones.filter(
            Q(ciudadano__primer_nombre__icontains=q) |
            Q(ciudadano__primer_apellido__icontains=q) |
            Q(ciudadano__numero_documento__icontains=q)
        )
    if estado_filter:
        atenciones = atenciones.filter(estado=estado_filter)
    if fecha_desde_filter:
        atenciones = atenciones.filter(fecha__gte=fecha_desde_filter)
    if fecha_hasta_filter:
        atenciones = atenciones.filter(fecha__lte=fecha_hasta_filter)

    total = atenciones.count()
    paginator = Paginator(atenciones, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    pvds_para_filtro = PuntoViveDigital.objects.filter(estado='A').order_by('nombre') if not es_admin_pvd_solo else None

    return render(request, 'modulo_puntos/lista_atenciones.html', {
        'atenciones': page_obj,
        'page_obj': page_obj,
        'total': total,
        'q': q,
        'estado_filter': estado_filter,
        'fecha_desde_filter': fecha_desde_filter,
        'fecha_hasta_filter': fecha_hasta_filter,
        'pvd_filter': pvd_filter,
        'pvds_para_filtro': pvds_para_filtro,
    })


@login_required(login_url='/login/')
def detalle_atencion(request, atencion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    atencion = get_object_or_404(
        Atencion.objects.select_related('ciudadano', 'punto_vive_digital', 'operador'),
        pk=atencion_id
    )
    if not pvd_permitido(request, atencion.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para ver atenciones de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')
    servicios = Servicio.objects.filter(atencion=atencion).order_by('pk')
    satisfaccion = Satisfaccion.objects.filter(atencion=atencion).first()

    return render(request, 'modulo_puntos/detalle_atencion.html', {
        'atencion': atencion,
        'servicios': servicios,
        'satisfaccion': satisfaccion,
    })


@login_required(login_url='/login/')
def cambiar_estado_atencion(request, atencion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    if request.method != 'POST':
        return redirect('modulo_puntos:lista_atenciones')

    atencion = get_object_or_404(Atencion, pk=atencion_id)
    if not pvd_permitido(request, atencion.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para modificar atenciones de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')
    nuevo_estado = request.POST.get('estado', '').strip()
    estados_validos = {'P', 'F', 'C'}

    etiquetas = {'P': 'Pendiente', 'F': 'Finalizada', 'C': 'Cancelada'}
    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado no válido.')
    elif nuevo_estado == 'C' and atencion.estado == 'P':
        messages.error(request, 'No se puede cancelar una atención pendiente directamente. Primero finalízala o cambia el motivo.')
    else:
        atencion.estado = nuevo_estado
        atencion.save(update_fields=['estado'])
        messages.success(request, f'Atención marcada como {etiquetas[nuevo_estado]}.')

    next_url = request.POST.get('next', '').strip()
    if next_url == 'lista':
        return redirect('modulo_puntos:lista_atenciones')
    return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)


@login_required(login_url='/login/')
def finalizar_servicio(request, atencion_id, servicio_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    if request.method != 'POST':
        return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion_id)

    servicio = get_object_or_404(
        Servicio.objects.select_related('atencion'), pk=servicio_id, atencion_id=atencion_id
    )
    if not pvd_permitido(request, servicio.atencion.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para modificar servicios de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')

    if servicio.estado == 'A':
        servicio.estado = 'F'
        servicio.save(update_fields=['estado'])
        registrar_auditoria(request, 'UPDATE', 'Servicio', servicio.pk,
                            f'Servicio finalizado: {servicio.nombre}')
        messages.success(request, f'Servicio "{servicio.nombre}" finalizado.')
    else:
        messages.warning(request, 'El servicio ya está finalizado o inactivo.')

    return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion_id)


@login_required(login_url='/login/')
def registrar_prestamo(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(request, 'Debes ingresar a un Punto Vive Digital antes de registrar un préstamo.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    pvd_activo = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first()
    if not pvd_activo:
        del request.session['pvd_activo_id']
        messages.warning(request, 'El Punto Vive Digital de tu sesión ya no está disponible. Selecciona otro.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    form = PrestamoRecursoForm(request.POST or None, pvd_id=pvd_id)
    if request.method == 'POST':
        if form.is_valid():
            try:
                prestamo = form.save()
                registrar_auditoria(request, 'CREATE', 'PrestamoRecurso', prestamo.pk,
                                    f'Nuevo préstamo: {prestamo.recurso}')
                messages.success(request, 'Préstamo registrado correctamente.')
                return redirect('modulo_puntos:registrar_recurso')
            except Exception as e:
                messages.error(request, f'Error al guardar en BD: {e}')
        else:
            if form.non_field_errors():
                messages.error(request, form.non_field_errors()[0])
            else:
                messages.error(request, 'Revisa los campos del formulario.')

    return render(request, 'modulo_puntos/registrar_prestamo.html', {'form': form})


@login_required(login_url='/login/')
def editar_prestamo(request, prestamo_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'inventario.editar_prestamo'):
        messages.error(request, 'No tienes permiso para editar préstamos.')
        return redirect('modulo_puntos:registrar_recurso')

    pvd_id = request.session.get('pvd_activo_id')
    prestamo = get_object_or_404(PrestamoRecurso.objects.select_related('recurso'), pk=prestamo_id)
    recurso_pvd_id = prestamo.recurso.punto_vive_digital_id if prestamo.recurso else None
    if not pvd_permitido(request, recurso_pvd_id):
        messages.error(request, 'No tienes permiso para editar préstamos de otro PVD.')
        return redirect('modulo_puntos:registrar_recurso')
    form = PrestamoRecursoForm(request.POST or None, instance=prestamo, pvd_id=pvd_id)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            registrar_auditoria(
                request, 'UPDATE', 'PrestamoRecurso', prestamo.pk,
                f'Préstamo editado: {prestamo.recurso} – entrega {prestamo.fecha_entrega}'
            )
            messages.success(request, 'Préstamo actualizado correctamente.')
            return redirect('modulo_puntos:registrar_recurso')  # lista_recursos
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/editar_prestamo.html', {
        'form': form,
        'prestamo': prestamo,
    })


@login_required(login_url='/login/')
def devolver_prestamo(request, prestamo_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    if request.method != 'POST':
        return redirect('modulo_puntos:registrar_recurso')

    from django.utils import timezone
    prestamo = get_object_or_404(PrestamoRecurso.objects.select_related('recurso'), pk=prestamo_id)
    recurso_pvd_id = prestamo.recurso.punto_vive_digital_id if prestamo.recurso else None
    if not pvd_permitido(request, recurso_pvd_id):
        messages.error(request, 'No tienes permiso para modificar préstamos de otro PVD.')
        return redirect('modulo_puntos:registrar_recurso')
    now = timezone.now()

    # Solo actuar si aún está en préstamo
    if prestamo.fecha_devolucion is None or prestamo.fecha_devolucion > now:
        prestamo.fecha_devolucion = now
        prestamo.save(update_fields=['fecha_devolucion'])
        registrar_auditoria(
            request, 'UPDATE', 'PrestamoRecurso', prestamo.pk,
            f'Devolución anticipada registrada: {prestamo.recurso} a las {now.strftime("%d/%m/%Y %H:%M")}'
        )
        messages.success(request, f'Devolución de "{prestamo.recurso}" registrada a las {now.strftime("%H:%M")}.')
    else:
        messages.info(request, 'Este préstamo ya figura como devuelto.')

    return redirect('modulo_puntos:registrar_recurso')


@login_required(login_url='/login/')
def lista_prestamos_global(request):
    """Vista global de todos los préstamos de ambos PVDs."""
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    from django.utils import timezone
    now = timezone.now()

    prestamos = (
        PrestamoRecurso.objects
        .select_related('recurso', 'recurso__punto_vive_digital', 'ciudadano')
        .order_by('-fecha_entrega')
    )

    q          = request.GET.get('q', '').strip()
    pvd_filter = request.GET.get('pvd_id', '').strip()
    estado     = request.GET.get('estado', '').strip()

    if usuario_necesita_seleccionar_pvd(request.user):
        # Admin PVD sólo ve préstamos de recursos de su PVD activo
        prestamos = prestamos.filter(recurso__punto_vive_digital_id=obtener_pvd_activo_id(request))
    elif pvd_filter:
        prestamos = prestamos.filter(recurso__punto_vive_digital_id=pvd_filter)

    if q:
        prestamos = prestamos.filter(
            Q(ciudadano__primer_nombre__icontains=q)
            | Q(ciudadano__primer_apellido__icontains=q)
            | Q(ciudadano__numero_documento__icontains=q)
            | Q(recurso__tipo__icontains=q)
            | Q(recurso__codigo__icontains=q)
        )
    if estado == 'activo':
        prestamos = prestamos.filter(Q(fecha_devolucion__isnull=True) | Q(fecha_devolucion__gt=now))
    elif estado == 'devuelto':
        prestamos = prestamos.filter(fecha_devolucion__lte=now)

    total = prestamos.count()
    pvds  = PuntoViveDigital.objects.filter(estado='A').order_by('nombre')

    paginator = Paginator(prestamos, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    for p in page_obj:
        p.ya_devuelto = p.fecha_devolucion is not None and p.fecha_devolucion <= now

    return render(request, 'modulo_puntos/lista_prestamos_global.html', {
        'prestamos':   page_obj,
        'page_obj':    page_obj,
        'total':       total,
        'pvds':        pvds,
        'q':           q,
        'pvd_filter':  pvd_filter,
        'estado':      estado,
    })


@login_required(login_url='/login/')
def lista_recursos(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    pvd = PuntoViveDigital.objects.filter(pk=pvd_id).first() if pvd_id else None

    from django.db.models import Prefetch
    from django.utils import timezone
    from itertools import groupby

    now = timezone.now()

    es_admin_pvd_solo = (
        not request.user.is_superuser
        and not usuario_es_admin_tic(request.user)
        and request.user.groups.filter(name='Administrador PVD').exists()
    )

    # Admin PVD solo ve su propio punto; Superusuario/Admin TIC ven todos los
    # PVD a la vez, así que para ellos se muestra a qué PVD pertenece cada recurso.
    mostrar_pvd_por_recurso = not es_admin_pvd_solo

    recursos_qs = Recurso.objects.all()
    if es_admin_pvd_solo:
        recursos_qs = recursos_qs.filter(punto_vive_digital_id=pvd_id) if pvd_id else recursos_qs.none()

    tipo_filter = request.GET.get('tipo', '').strip()
    q_filter = request.GET.get('q', '').strip()
    pvd_filter = request.GET.get('pvd_id', '').strip() if mostrar_pvd_por_recurso else ''

    tipos_disponibles = sorted(recursos_qs.values_list('tipo', flat=True).distinct())
    pvds_disponibles = (
        PuntoViveDigital.objects.filter(estado='A').order_by('nombre')
        if mostrar_pvd_por_recurso else []
    )

    display_qs = recursos_qs
    if tipo_filter:
        display_qs = display_qs.filter(tipo=tipo_filter)
    if pvd_filter:
        display_qs = display_qs.filter(punto_vive_digital_id=pvd_filter)
    if q_filter:
        display_qs = display_qs.filter(
            Q(tipo__icontains=q_filter) | Q(codigo__icontains=q_filter)
        )

    recursos = list(
        display_qs
          .select_related('punto_vive_digital')
          .annotate(total_prestamos=Count('prestamorecurso'))
          .prefetch_related(
              Prefetch('prestamorecurso_set',
                       queryset=PrestamoRecurso.objects.order_by('-fecha_entrega'),
                       to_attr='todos_los_prestamos')
          )
          .order_by('tipo', 'punto_vive_digital__nombre', 'codigo')
    )

    for recurso in recursos:
        for p in recurso.todos_los_prestamos:
            p.ya_devuelto = p.fecha_devolucion is not None and p.fecha_devolucion <= now
        ultimo = recurso.todos_los_prestamos[0] if recurso.todos_los_prestamos else None
        recurso.prestado_ahora = ultimo is not None and not ultimo.ya_devuelto if ultimo else False

    prestados_count = sum(1 for r in recursos if r.prestado_ahora)
    disponibles_count = len(recursos) - prestados_count

    total_filtrados = len(recursos)
    # Página grande: cada tarjeta ahora es compacta y se agrupan por tipo,
    # así un mismo tipo (ej. "Portátil") no queda partido entre dos páginas.
    paginator = Paginator(recursos, 200)
    page = request.GET.get('page')
    recursos_page = paginator.get_page(page)

    grupos = []
    for tipo, items in groupby(recursos_page, key=lambda r: r.tipo):
        items = list(items)
        grupos.append({
            'tipo': tipo,
            'items': items,
            'total': len(items),
            'disponibles': sum(1 for r in items if not r.prestado_ahora),
            'prestados': sum(1 for r in items if r.prestado_ahora),
        })

    return render(request, 'modulo_puntos/lista_recursos.html', {
        'recursos': recursos_page,
        'grupos': grupos,
        'pvd': pvd,
        'mostrar_pvd_por_recurso': mostrar_pvd_por_recurso,
        'total_recursos': total_filtrados,
        'prestados_count': prestados_count,
        'disponibles_count': disponibles_count,
        'tipos_disponibles': tipos_disponibles,
        'tipo_filter': tipo_filter,
        'q_filter': q_filter,
        'pvds_disponibles': pvds_disponibles,
        'pvd_filter': pvd_filter,
        'puede_eliminar_recursos': tiene_permiso(request.user, 'inventario.eliminar_recurso'),
    })


@login_required(login_url='/login/')
def crear_recurso(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para registrar recursos.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(request, 'Debes ingresar a un Punto Vive Digital antes de agregar un recurso.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    pvd = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first()
    if not pvd:
        del request.session['pvd_activo_id']
        messages.warning(request, 'El Punto Vive Digital de tu sesión ya no está disponible. Selecciona otro.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    form = RecursoForm(request.POST or None, pvd_id=pvd.pk)
    if request.method == 'POST':
        if form.is_valid():
            recurso = form.save(commit=False)
            recurso.punto_vive_digital = pvd
            recurso.save()
            registrar_auditoria(request, 'CREATE', 'Recurso', recurso.pk, f'Nuevo recurso: {recurso.tipo} — PVD: {pvd}')
            messages.success(request, f'Recurso "{recurso.tipo}" registrado correctamente.')
            return redirect('modulo_puntos:registrar_recurso')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/registrar_recurso.html', {'form': form, 'pvd': pvd})


@login_required(login_url='/login/')
def registrar_satisfaccion(request, atencion_id=None):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    # PVD en sesión es obligatorio para registrar satisfacción
    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(
            request,
            'Debes ingresar a un Punto Vive Digital antes de registrar una encuesta de satisfacción.'
        )
        return redirect('modulo_puntos:seleccionar_pvd_view')

    pvd_activo = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first()
    if not pvd_activo:
        del request.session['pvd_activo_id']
        messages.warning(request, 'El Punto Vive Digital de tu sesión ya no está disponible. Selecciona otro.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    atencion = None
    if atencion_id:
        atencion = get_object_or_404(
            Atencion.objects.select_related('ciudadano', 'punto_vive_digital'),
            pk=atencion_id,
        )
        # Verificar que la atención pertenece al PVD activo
        if atencion.punto_vive_digital_id != pvd_activo.pk:
            messages.error(request, 'Esa atención no pertenece al Punto Vive Digital activo en tu sesión.')
            return redirect('modulo_puntos:panel_control')
        # Verificar que la atención tiene ciudadano
        if not atencion.ciudadano_id:
            messages.error(request, 'No se puede registrar satisfacción: la atención no tiene ciudadano asignado.')
            return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
        if Satisfaccion.objects.filter(atencion=atencion).exists():
            messages.warning(request, 'Esta atención ya tiene una encuesta de satisfacción registrada.')
            return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)

    from django.utils import timezone
    initial = {'fecha': timezone.now().strftime('%Y-%m-%dT%H:%M')}
    if atencion:
        initial['atencion'] = atencion.pk

    form = SatisfaccionForm(request.POST or None, initial=initial, pvd_id=pvd_id)
    if atencion:
        form.fields['atencion'].widget = forms.HiddenInput()
        form.fields['atencion'].required = False   # oculto: ya viene fijo por atencion_id

    if request.method == 'POST':
        if form.is_valid():
            try:
                satisfaccion = form.save(commit=False)
                # Si viene de detalle de atención, forzar la atención correcta
                if atencion:
                    satisfaccion.atencion = atencion
                satisfaccion.save()
                registrar_auditoria(
                    request, 'CREATE', 'Satisfaccion', satisfaccion.pk,
                    f'Encuesta registrada en {pvd_activo.nombre}'
                )
                messages.success(request, 'Encuesta de satisfacción registrada correctamente.')
                if atencion:
                    return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
                return redirect('modulo_puntos:panel_control')
            except Exception as e:
                logger.error('Error al guardar satisfacción: %s', e, exc_info=True)
                messages.error(request, f'Error al guardar en BD: {e}')
        else:
            messages.error(request, 'No se pudo guardar la satisfacción. Revisa los datos ingresados.')

    return render(request, 'modulo_puntos/registrar_satisfaccion.html', {
        'form': form,
        'atencion': atencion,
        'pvd_activo': pvd_activo,
    })


@login_required(login_url='/login/')
def gestionar_servicios_pvd(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    user = request.user
    es_admin_pvd_solo = (
        not user.is_superuser
        and not usuario_es_admin_tic(user)
        and user.groups.filter(name='Administrador PVD').exists()
    )

    q           = request.GET.get('q', '').strip()
    tipo_filter = request.GET.get('tipo', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    servicios = (
        Servicio.objects
        .select_related('atencion', 'atencion__ciudadano', 'atencion__punto_vive_digital', 'recurso')
        .order_by('-atencion__fecha', '-pk')
    )

    if es_admin_pvd_solo:
        pvd_id = request.session.get('pvd_activo_id')
        servicios = servicios.filter(atencion__punto_vive_digital_id=pvd_id) if pvd_id else servicios.none()

    if q:
        servicios = servicios.filter(
            Q(atencion__ciudadano__primer_nombre__icontains=q) |
            Q(atencion__ciudadano__primer_apellido__icontains=q) |
            Q(atencion__ciudadano__numero_documento__icontains=q)
        )
    if tipo_filter:
        servicios = servicios.filter(tipo=tipo_filter)
    if fecha_desde:
        servicios = servicios.filter(atencion__fecha__gte=fecha_desde)
    if fecha_hasta:
        servicios = servicios.filter(atencion__fecha__lte=fecha_hasta)

    from .forms import TIPO_SERVICIO_CHOICES
    tipos = [c for c in TIPO_SERVICIO_CHOICES if c[0]]

    total = servicios.count()
    paginator = Paginator(servicios, 25)
    page = request.GET.get('page')
    servicios_page = paginator.get_page(page)

    return render(request, 'modulo_puntos/gestionar_servicios_pvd.html', {
        'servicios': servicios_page,
        'q': q,
        'tipo_filter': tipo_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos': tipos,
        'total': total,
    })


@login_required(login_url='/login/')
def registrar_servicio(request, atencion_id=None):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    atencion = get_object_or_404(Atencion, pk=atencion_id) if atencion_id else None
    if atencion and not pvd_permitido(request, atencion.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para registrar servicios en atenciones de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')

    pvd_id = request.session.get('pvd_activo_id')
    if pvd_id:
        recursos_pvd = Recurso.objects.filter(punto_vive_digital_id=pvd_id, estado='A').order_by('tipo', 'codigo')
        salas_pvd = Sala.objects.filter(punto_vive_digital_id=pvd_id, estado='A').order_by('nombre')
    else:
        recursos_pvd = Recurso.objects.filter(estado='A').order_by('tipo', 'codigo')
        salas_pvd = Sala.objects.filter(estado='A').order_by('nombre')

    initial = {'atencion': atencion} if atencion else {}
    form = ServicioForm(request.POST or None, initial=initial, recursos_pvd=recursos_pvd, salas_pvd=salas_pvd)

    if atencion:
        form.fields['atencion'].widget.attrs['disabled'] = True
        form.fields['atencion'].required = False

    if request.method == 'POST':
        if form.is_valid():
            try:
                servicio = form.save(commit=False)
                if atencion:
                    servicio.atencion = atencion
                servicio.save()
                messages.success(request, 'Servicio registrado correctamente.')
                if servicio.requiere_sala == 'S' and servicio.sala_id:
                    url = reverse('modulo_puntos:crear_habilitacion') + f'?sala_id={servicio.sala_id}'
                    if atencion:
                        url += f'&atencion_id={atencion.pk}'
                    return redirect(url)
                if atencion:
                    return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
                return redirect('modulo_puntos:panel_control')
            except Exception as e:
                messages.error(request, f'Error al guardar en BD: {e}')
        else:
            messages.error(request, 'No se pudo guardar el servicio. Revisa los datos ingresados.')

    return render(request, 'modulo_puntos/registrar_servicio.html', {
        'form': form,
        'atencion': atencion,
    })


# ── EDITAR SERVICIO ────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def editar_servicio(request, servicio_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    servicio = get_object_or_404(Servicio.objects.select_related('atencion'), pk=servicio_id)
    atencion = servicio.atencion
    if not pvd_permitido(request, atencion.punto_vive_digital_id if atencion else None):
        messages.error(request, 'No tienes permiso para editar servicios de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')

    pvd_id = request.session.get('pvd_activo_id')
    if pvd_id:
        recursos_pvd = Recurso.objects.filter(punto_vive_digital_id=pvd_id, estado='A').order_by('tipo', 'codigo')
        salas_pvd = Sala.objects.filter(punto_vive_digital_id=pvd_id, estado='A').order_by('nombre')
    else:
        recursos_pvd = Recurso.objects.filter(estado='A').order_by('tipo', 'codigo')
        salas_pvd = Sala.objects.filter(estado='A').order_by('nombre')

    form = ServicioForm(request.POST or None, instance=servicio, recursos_pvd=recursos_pvd, salas_pvd=salas_pvd)
    form.fields['atencion'].widget.attrs['disabled'] = True
    form.fields['atencion'].required = False

    if request.method == 'POST':
        if form.is_valid():
            s = form.save(commit=False)
            s.atencion = atencion
            s.save()
            registrar_auditoria(request, 'UPDATE', 'Servicio', servicio.pk, f'Servicio editado: {servicio.nombre}')
            messages.success(request, 'Servicio actualizado correctamente.')
            return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/registrar_servicio.html', {
        'form': form,
        'atencion': atencion,
        'editando': True,
        'servicio': servicio,
    })


# ── ELIMINAR SERVICIO ─────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def eliminar_servicio(request, servicio_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    servicio = get_object_or_404(Servicio.objects.select_related('atencion'), pk=servicio_id)
    atencion_id = servicio.atencion_id
    if not pvd_permitido(request, servicio.atencion.punto_vive_digital_id if servicio.atencion else None):
        messages.error(request, 'No tienes permiso para eliminar servicios de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')

    if request.method == 'POST':
        registrar_auditoria(request, 'DELETE', 'Servicio', servicio.pk,
                            f'Servicio eliminado: {servicio.nombre} — Atención #{atencion_id}')
        servicio.delete()
        messages.success(request, 'Servicio eliminado correctamente.')

    return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion_id)


# ── EDITAR/ELIMINAR SATISFACCIÓN ───────────────────────────────────────────────

@login_required(login_url='/login/')
def editar_satisfaccion(request, satisfaccion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    satisfaccion = get_object_or_404(Satisfaccion.objects.select_related('atencion'), pk=satisfaccion_id)
    atencion = satisfaccion.atencion
    if not pvd_permitido(request, atencion.punto_vive_digital_id if atencion else None):
        messages.error(request, 'No tienes permiso para editar encuestas de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')

    form = SatisfaccionForm(request.POST or None, instance=satisfaccion)
    form.fields['atencion'].widget = forms.HiddenInput()

    if request.method == 'POST':
        if form.is_valid():
            s = form.save(commit=False)
            s.atencion = atencion
            s.save()
            registrar_auditoria(request, 'UPDATE', 'Satisfaccion', satisfaccion.pk, f'Satisfacción editada — Atención #{atencion.pk}')
            messages.success(request, 'Encuesta de satisfacción actualizada.')
            return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion.pk)
        messages.error(request, 'Revisa los datos.')

    return render(request, 'modulo_puntos/registrar_satisfaccion.html', {
        'form': form,
        'atencion': atencion,
        'editando': True,
    })


@login_required(login_url='/login/')
def eliminar_satisfaccion(request, satisfaccion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    satisfaccion = get_object_or_404(Satisfaccion.objects.select_related('atencion'), pk=satisfaccion_id)
    atencion_id = satisfaccion.atencion_id
    if not pvd_permitido(request, satisfaccion.atencion.punto_vive_digital_id if satisfaccion.atencion else None):
        messages.error(request, 'No tienes permiso para eliminar encuestas de otro PVD.')
        return redirect('modulo_puntos:lista_atenciones')

    if request.method == 'POST':
        registrar_auditoria(request, 'DELETE', 'Satisfaccion', satisfaccion.pk, f'Satisfacción eliminada — Atención #{atencion_id}')
        satisfaccion.delete()
        messages.success(request, 'Encuesta de satisfacción eliminada.')

    return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion_id)


# ── EDITAR EVIDENCIA ───────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def editar_evidencia(request, evidencia_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    evidencia = get_object_or_404(Evidencia, pk=evidencia_id)
    pvd_id = request.session.get('pvd_activo_id')
    if pvd_id and not request.user.is_superuser and not usuario_es_admin_tic(request.user):
        if evidencia.punto_vive_digital_id != pvd_id:
            messages.error(request, 'No tienes permisos para editar esta evidencia.')
            return redirect('modulo_puntos:lista_evidencias')

    form = EvidenciaForm(request.POST or None, request.FILES or None, instance=evidencia)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            registrar_auditoria(request, 'UPDATE', 'Evidencia', evidencia.pk, f'Evidencia editada: {evidencia.titulo}')
            messages.success(request, f'Evidencia "{evidencia.titulo}" actualizada.')
            return redirect('modulo_puntos:lista_evidencias')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/crear_evidencia.html', {
        'form': form,
        'pvd': evidencia.punto_vive_digital,
        'editando': True,
        'evidencia': evidencia,
    })


# ── EDITAR RECURSO ─────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def editar_recurso(request, recurso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    recurso = get_object_or_404(Recurso, pk=recurso_id)
    if not pvd_permitido(request, recurso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para editar recursos de otro PVD.')
        return redirect('modulo_puntos:registrar_recurso')
    form = RecursoForm(request.POST or None, instance=recurso)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            registrar_auditoria(request, 'UPDATE', 'Recurso', recurso.pk, f'Recurso editado: {recurso.tipo} ({recurso.codigo})')
            messages.success(request, f'Recurso "{recurso.tipo}" actualizado correctamente.')
            return redirect('modulo_puntos:registrar_recurso')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/registrar_recurso.html', {
        'form': form,
        'editando': True,
        'recurso': recurso,
    })


# ── ELIMINAR RECURSO ──────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def eliminar_recurso(request, recurso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'inventario.eliminar_recurso'):
        messages.error(request, 'No tienes permiso para eliminar recursos del inventario.')
        return redirect('modulo_puntos:registrar_recurso')

    recurso = get_object_or_404(Recurso, pk=recurso_id)
    if not pvd_permitido(request, recurso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para eliminar recursos de otro PVD.')
        return redirect('modulo_puntos:registrar_recurso')

    if request.method == 'POST':
        prestamos_activos = PrestamoRecurso.objects.filter(recurso=recurso, fecha_devolucion=None).count()
        if prestamos_activos:
            messages.error(request, f'No se puede eliminar: el recurso tiene {prestamos_activos} préstamo(s) activo(s).')
            return redirect('modulo_puntos:registrar_recurso')
        registrar_auditoria(request, 'DELETE', 'Recurso', recurso.pk,
                            f'Recurso eliminado: {recurso.tipo} ({recurso.codigo or "sin código"})')
        recurso.delete()
        messages.success(request, 'Recurso eliminado correctamente.')

    return redirect('modulo_puntos:registrar_recurso')


# ── EXPORTACIÓN XLSX ───────────────────────────────────────────────────────────

def _xlsx_response(filename_base, wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}_{fecha_actual}.xlsx"'
    return response


def _estilizar_hoja(ws, headers, filas):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    cab_fondo = PatternFill(start_color='0B1220', end_color='0B1220', fill_type='solid')
    cab_fuente = Font(bold=True, color='FFFFFF', size=10)
    par_fondo = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    borde_cab = Border(
        left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='medium', color='4A90D9'),
    )
    borde_dato = Border(
        left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'),
    )
    for col, texto in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=texto)
        c.font = cab_fuente; c.fill = cab_fondo; c.border = borde_cab
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32
    for ri, fila in enumerate(filas, 2):
        for ci, valor in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=valor)
            c.alignment = Alignment(vertical='center', wrap_text=False)
            c.border = borde_dato
            if ri % 2 == 0:
                c.fill = par_fondo
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _crear_hoja(titulo, headers, filas):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = titulo[:31]
    _estilizar_hoja(wb.active, headers, filas)
    return wb


def _agregar_hoja(wb, titulo, headers, filas):
    ws = wb.create_sheet(titulo[:31])
    _estilizar_hoja(ws, headers, filas)
    return ws


@login_required(login_url='/login/')
def exportar_atenciones_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID Atención', 'Punto Vive Digital', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Estado',
        'Documento Ciudadano', 'Nombre Completo', 'Género', 'Etnia',
        'Discapacidad', 'Detalle Discapacidad', 'Barrio', 'Dirección',
        'Vereda / Corregimiento', 'Admin PVD a Cargo', 'Observaciones',
    ]
    estado_dict = dict(Atencion.ESTADO_CHOICES)
    qs = Atencion.objects.select_related('ciudadano', 'operador', 'punto_vive_digital').order_by('-fecha', '-hora_inicio')
    if pvd_id:
        qs = qs.filter(punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    filas = []
    for a in qs:
        c = a.ciudadano
        if c:
            doc = c.numero_documento or ''
            nom = f"{c.primer_nombre or ''} {c.primer_apellido or ''}".strip()
            gen = c.get_genero_display()
            etnia = c.etnia or ''
            discap = 'Sí' if c.tiene_discapacidad else 'No'
            desc_discap = c.descripcion_discapacidad or ''
            barrio = c.barrio or ''
            direccion = c.direccion or ''
            rural = c.zona_rural or ''
        else:
            doc = nom = gen = etnia = discap = desc_discap = barrio = direccion = rural = ''
        operador = a.operador.get_full_name() or a.operador.username if a.operador else ''
        pvd_nombre = a.punto_vive_digital.nombre if a.punto_vive_digital else ''
        filas.append([
            a.pk,
            pvd_nombre,
            str(a.fecha),
            str(a.hora_inicio),
            str(a.hora_fin) if a.hora_fin else '',
            estado_dict.get(a.estado, a.estado),
            doc, nom, gen, etnia, discap, desc_discap,
            barrio, direccion, rural, operador,
            a.observaciones or '',
        ])

    wb = _crear_hoja('Atenciones', headers, filas)
    return _xlsx_response('Reporte_Atenciones_PVD', wb)


@login_required(login_url='/login/')
@login_required(login_url='/login/')
def exportar_ciudadanos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    headers = [
        'ID', 'Punto Vive Digital', 'Tipo Documento', 'Número Documento',
        'Primer Nombre', 'Segundo Nombre', 'Primer Apellido', 'Segundo Apellido',
        'Fecha Nacimiento', 'Género', 'Etnia', 'Nivel Educativo', 'Ocupación',
        'Discapacidad', 'Descripción Discapacidad',
        'Dirección', 'Barrio', 'Zona Rural', 'Estrato', 'Estado', 'Email', 'Teléfono',
    ]
    qs = Ciudadano.objects.select_related('punto_vive_digital').order_by('-pk')
    if pvd_id:
        qs = qs.filter(punto_vive_digital_id=pvd_id)
    filas = []
    for c in qs:
        filas.append([
            c.pk,
            c.punto_vive_digital.nombre if c.punto_vive_digital else '',
            c.tipo_documento or '',
            c.numero_documento or '',
            c.primer_nombre or '',
            c.segundo_nombre or '',
            c.primer_apellido or '',
            c.segundo_apellido or '',
            str(c.fecha_nacimiento) if c.fecha_nacimiento else '',
            c.get_genero_display(),
            c.etnia or '',
            c.nivel_educativo or '',
            c.ocupacion or '',
            'Sí' if c.tiene_discapacidad else 'No',
            c.descripcion_discapacidad or '',
            c.direccion or '',
            c.barrio or '',
            c.zona_rural or '',
            c.estrato,
            c.get_estado_display(),
            c.correo or '',
            c.telefono or '',
        ])

    wb = _crear_hoja('Ciudadanos', headers, filas)
    return _xlsx_response('Reporte_Ciudadanos_PVD', wb)


@login_required(login_url='/login/')
def exportar_servicios_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID Servicio', 'Punto Vive Digital', 'ID Atención', 'Fecha Atención',
        'Documento Ciudadano', 'Nombre Ciudadano',
        'Nombre Servicio', 'Descripción', 'Tipo Servicio', 'Requiere Equipo', 'Recurso Utilizado', 'Estado',
    ]
    qs = Servicio.objects.select_related(
        'atencion', 'atencion__ciudadano', 'atencion__punto_vive_digital', 'recurso'
    ).order_by('-pk')
    if pvd_id:
        qs = qs.filter(atencion__punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs = qs.filter(atencion__fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(atencion__fecha__lte=fecha_hasta)
    filas = []
    for s in qs:
        atn = s.atencion
        fecha_atn = str(atn.fecha) if atn else ''
        pvd_nombre = atn.punto_vive_digital.nombre if atn and atn.punto_vive_digital else ''
        doc = nom = ''
        if atn and atn.ciudadano:
            doc = atn.ciudadano.numero_documento or ''
            nom = f"{atn.ciudadano.primer_nombre or ''} {atn.ciudadano.primer_apellido or ''}".strip()
        filas.append([
            s.pk,
            pvd_nombre,
            atn.pk if atn else '',
            fecha_atn,
            doc, nom,
            s.nombre,
            s.descripcion or '',
            s.tipo,
            s.get_requiere_equipo_display(),
            str(s.recurso) if s.recurso else '',
            s.get_estado_display(),
        ])

    wb = _crear_hoja('Servicios', headers, filas)
    return _xlsx_response('Reporte_Servicios_PVD', wb)


@login_required(login_url='/login/')
def exportar_satisfaccion_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    headers = [
        'ID', 'Punto Vive Digital', 'ID Atención', 'Fecha Atención', 'Estado Atención',
        'Documento Ciudadano', 'Nombre Ciudadano',
        '¿Tiempo de espera para ser atendido?', '¿Atención brindada por el servidor público?',
        '¿Quedó satisfecho con la prestación del servicio?', '¿La información recibida fue?',
        '¿Comodidad y limpieza de las instalaciones?', 'Puntaje promedio (1-5)',
        'Comentario', 'Fecha de Registro',
    ]
    estado_atn = dict(Atencion.ESTADO_CHOICES)
    qs = Satisfaccion.objects.select_related('atencion', 'atencion__ciudadano', 'atencion__punto_vive_digital').order_by('-pk')
    if pvd_id:
        qs = qs.filter(atencion__punto_vive_digital_id=pvd_id)
    filas = []
    for sat in qs:
        atn = sat.atencion
        doc = nom = fecha_atn = est_atn = pvd_nombre = ''
        if atn:
            fecha_atn = str(atn.fecha)
            est_atn = estado_atn.get(atn.estado, atn.estado)
            pvd_nombre = atn.punto_vive_digital.nombre if atn.punto_vive_digital else ''
            if atn.ciudadano:
                doc = atn.ciudadano.numero_documento or ''
                nom = f"{atn.ciudadano.primer_nombre or ''} {atn.ciudadano.primer_apellido or ''}".strip()
        promedio = sat.puntaje_promedio
        filas.append([
            sat.pk,
            pvd_nombre,
            atn.pk if atn else '',
            fecha_atn, est_atn,
            doc, nom,
            sat.get_tiempo_espera_display(),
            sat.get_atencion_servidor_display(),
            sat.get_satisfaccion_servicio_display(),
            sat.get_informacion_recibida_display(),
            sat.get_comodidad_instalaciones_display(),
            round(promedio, 1) if promedio is not None else '',
            sat.comentario or '',
            str(sat.fecha),
        ])

    wb = _crear_hoja('Satisfaccion', headers, filas)
    return _xlsx_response('Reporte_Satisfaccion_PVD', wb)


@login_required(login_url='/login/')
def exportar_prestamos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID Préstamo', 'Punto Vive Digital', 'Tipo Recurso', 'Código Recurso',
        'Fecha de Entrega', 'Fecha de Devolución', 'Observaciones', 'Estado',
    ]
    qs = PrestamoRecurso.objects.select_related('recurso', 'recurso__punto_vive_digital').order_by('-pk')
    if pvd_id:
        qs = qs.filter(recurso__punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs = qs.filter(fecha_entrega__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_entrega__lte=fecha_hasta)
    filas = []
    for p in qs:
        tipo = p.recurso.tipo if p.recurso else ''
        codigo = p.recurso.codigo or '' if p.recurso else ''
        pvd_nombre = p.recurso.punto_vive_digital.nombre if p.recurso and p.recurso.punto_vive_digital else ''
        estado = 'Activo (sin devolución)' if not p.fecha_devolucion else 'Devuelto'
        filas.append([
            p.pk,
            pvd_nombre,
            tipo,
            codigo,
            str(p.fecha_entrega),
            str(p.fecha_devolucion) if p.fecha_devolucion else '',
            p.observaciones or '',
            estado,
        ])

    wb = _crear_hoja('Prestamos', headers, filas)
    return _xlsx_response('Reporte_Prestamos_PVD', wb)


@login_required(login_url='/login/')
def exportar_cursos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    # Hoja 1 – Cursos / Talleres
    qs_cursos = (
        Curso.objects
        .select_related('punto_vive_digital', 'registrado_por')
        .order_by('-fecha_inicio')
    )
    if pvd_id:
        qs_cursos = qs_cursos.filter(punto_vive_digital_id=pvd_id)
    if fecha_desde:
        qs_cursos = qs_cursos.filter(fecha_inicio__gte=fecha_desde)
    if fecha_hasta:
        qs_cursos = qs_cursos.filter(fecha_inicio__lte=fecha_hasta)

    headers_cursos = [
        'ID', 'Punto Vive Digital', 'Nombre del Curso/Taller',
        'Modalidad', 'Población Objetivo', 'Fecha Inicio', 'Fecha Fin',
        'Estado', 'Total Inscritos', 'Registrado por',
    ]
    filas_cursos = []
    for c in qs_cursos:
        filas_cursos.append([
            c.pk,
            c.punto_vive_digital.nombre if c.punto_vive_digital else '',
            c.nombre,
            c.get_modalidad_display(),
            c.poblacion_objetivo or '',
            str(c.fecha_inicio),
            str(c.fecha_fin) if c.fecha_fin else '',
            c.get_estado_display(),
            c.total_inscritos(),
            c.registrado_por.get_full_name() or c.registrado_por.username if c.registrado_por else '',
        ])

    # Hoja 2 – Inscripciones
    qs_inscripciones = (
        InscripcionCurso.objects
        .select_related('curso', 'curso__punto_vive_digital', 'ciudadano', 'registrado_por')
        .order_by('curso__nombre', 'ciudadano__primer_apellido')
    )
    if pvd_id:
        qs_inscripciones = qs_inscripciones.filter(curso__punto_vive_digital_id=pvd_id)

    headers_inscripciones = [
        'ID Inscripción', 'Punto Vive Digital', 'Curso / Taller',
        'Documento Ciudadano', 'Nombre Ciudadano',
        'Estado Inscripción', 'Fecha Inscripción',
    ]
    filas_inscripciones = []
    for i in qs_inscripciones:
        ciu = i.ciudadano
        filas_inscripciones.append([
            i.pk,
            i.curso.punto_vive_digital.nombre if i.curso.punto_vive_digital else '',
            i.curso.nombre,
            ciu.numero_documento or '' if ciu else '',
            f"{ciu.primer_nombre or ''} {ciu.primer_apellido or ''}".strip() if ciu else '',
            i.get_estado_display(),
            str(i.fecha_inscripcion.date()) if i.fecha_inscripcion else '',
        ])

    # Hoja 3 – Asistencias por sesión
    qs_asistencias = (
        AsistenciaSesion.objects
        .select_related(
            'sesion', 'sesion__curso', 'sesion__curso__punto_vive_digital',
            'ciudadano',
        )
        .order_by('sesion__curso__nombre', 'sesion__numero_sesion', 'ciudadano__primer_apellido')
    )
    if pvd_id:
        qs_asistencias = qs_asistencias.filter(sesion__curso__punto_vive_digital_id=pvd_id)

    headers_asistencias = [
        'Punto Vive Digital', 'Curso / Taller', 'N° Sesión',
        'Fecha Sesión', 'Tema', 'Documento Ciudadano', 'Nombre Ciudadano', 'Asistió',
    ]
    filas_asistencias = []
    for a in qs_asistencias:
        ciu = a.ciudadano
        sesion = a.sesion
        filas_asistencias.append([
            sesion.curso.punto_vive_digital.nombre if sesion.curso.punto_vive_digital else '',
            sesion.curso.nombre,
            sesion.numero_sesion,
            str(sesion.fecha),
            sesion.tema,
            ciu.numero_documento or '' if ciu else '',
            f"{ciu.primer_nombre or ''} {ciu.primer_apellido or ''}".strip() if ciu else '',
            'Sí' if a.asistio else 'No',
        ])

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _agregar_hoja(wb, 'Cursos y Talleres', headers_cursos, filas_cursos)
    _agregar_hoja(wb, 'Inscripciones', headers_inscripciones, filas_inscripciones)
    _agregar_hoja(wb, 'Asistencias', headers_asistencias, filas_asistencias)
    return _xlsx_response('Reporte_Cursos_PVD', wb)


@login_required(login_url='/login/')
def exportar_mantenimientos_csv(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = obtener_pvd_activo_id(request)
    if usuario_necesita_seleccionar_pvd(request.user) and not pvd_id:
        # Fail-closed: sin PVD activo, un Admin PVD no puede exportar
        # (antes esto exportaba los datos de todos los PVD sin filtrar).
        messages.warning(request, 'Selecciona un Punto Vive Digital antes de exportar.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    fecha_desde_m = request.GET.get('fecha_desde', '').strip()
    fecha_hasta_m = request.GET.get('fecha_hasta', '').strip()

    headers = [
        'ID', 'Punto Vive Digital', 'Tipo', 'Fecha',
        'Equipos Intervenidos', 'Descripción del Trabajo',
        'Hallazgos', 'Acciones / Recomendaciones', 'Registrado por',
    ]
    qs = (
        MantenimientoEquipo.objects
        .select_related('punto_vive_digital', 'realizado_por')
        .order_by('-fecha')
    )
    if pvd_id:
        qs = qs.filter(punto_vive_digital_id=pvd_id)
    if fecha_desde_m:
        qs = qs.filter(fecha__gte=fecha_desde_m)
    if fecha_hasta_m:
        qs = qs.filter(fecha__lte=fecha_hasta_m)

    filas = []
    for m in qs:
        filas.append([
            m.pk,
            m.punto_vive_digital.nombre if m.punto_vive_digital else '',
            m.get_tipo_display(),
            str(m.fecha),
            m.equipos_intervenidos,
            m.descripcion,
            m.hallazgos or '',
            m.acciones or '',
            m.realizado_por.get_full_name() or m.realizado_por.username if m.realizado_por else '',
        ])

    wb = _crear_hoja('Mantenimientos', headers, filas)
    return _xlsx_response('Reporte_Mantenimientos_PVD', wb)


# ── EVIDENCIAS ────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def lista_evidencias(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')

    es_admin_pvd_solo = (
        not request.user.is_superuser
        and not usuario_es_admin_tic(request.user)
        and request.user.groups.filter(name='Administrador PVD').exists()
    )

    pvd_id = request.session.get('pvd_activo_id')
    pvd = PuntoViveDigital.objects.filter(pk=pvd_id).first() if pvd_id else None

    qs = Evidencia.objects.select_related('punto_vive_digital', 'registrado_por').order_by('-fecha', '-fecha_registro')
    if es_admin_pvd_solo and pvd:
        qs = qs.filter(punto_vive_digital=pvd)

    categoria = request.GET.get('categoria', '')
    if categoria:
        qs = qs.filter(categoria=categoria)

    paginator = Paginator(qs, 12)
    page = request.GET.get('page')
    evidencias = paginator.get_page(page)

    return render(request, 'modulo_puntos/lista_evidencias.html', {
        'evidencias': evidencias,
        'pvd': pvd,
        'categoria_activa': categoria,
        'categorias': Evidencia.CATEGORIA_CHOICES,
        'total': qs.count(),
        'es_admin_pvd_solo': es_admin_pvd_solo,
    })


@login_required(login_url='/login/')
def crear_evidencia(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para registrar evidencias.')
        return redirect('modulo_puntos:panel_control')
    # Solo el Admin PVD y el Superusuario pueden subir evidencias
    _es_admin_pvd = (
        not request.user.is_superuser
        and not usuario_es_admin_tic(request.user)
        and request.user.groups.filter(name='Administrador PVD').exists()
    )
    if not _es_admin_pvd and not request.user.is_superuser:
        messages.error(request, 'Solo el Administrador PVD o el Superusuario pueden registrar evidencias de trabajo.')
        return redirect('modulo_puntos:lista_evidencias')

    pvd_id = request.session.get('pvd_activo_id')
    pvd = PuntoViveDigital.objects.filter(pk=pvd_id, estado='A').first() if pvd_id else None

    if not pvd:
        messages.warning(request, 'Debes seleccionar un Punto Vive Digital antes de registrar evidencias.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    form = EvidenciaForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            evidencia = form.save(commit=False)
            evidencia.punto_vive_digital = pvd
            evidencia.registrado_por = request.user
            evidencia.save()
            registrar_auditoria(request, 'CREATE', 'Evidencia', evidencia.pk, f'Nueva evidencia: {evidencia.titulo} — PVD: {pvd}')
            messages.success(request, f'Evidencia "{evidencia.titulo}" registrada correctamente.')
            return redirect('modulo_puntos:lista_evidencias')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/crear_evidencia.html', {'form': form, 'pvd': pvd})


@login_required(login_url='/login/')
def eliminar_evidencia(request, evidencia_id):
    messages.error(request, 'Las evidencias no pueden ser eliminadas. Contacta al Administrador TIC si necesitas gestionar esta evidencia.')
    return redirect('modulo_puntos:lista_evidencias')


@login_required(login_url='/login/')
def servir_evidencia(request, evidencia_id):
    """
    Sirve la imagen de una Evidencia solo a usuarios con sesión iniciada
    (y con permiso sobre el PVD dueño de la evidencia). El archivo en sí
    vive fuera del alcance público: en producción Nginx no expone /media/
    directamente, así que este es el único camino para verlo.
    """
    if not usuario_puede_usar_modulos_pvd(request.user):
        raise Http404
    evidencia = get_object_or_404(Evidencia, pk=evidencia_id)
    if not pvd_permitido(request, evidencia.punto_vive_digital_id):
        raise Http404
    if not evidencia.imagen:
        raise Http404

    content_type = mimetypes.guess_type(evidencia.imagen.name)[0] or 'application/octet-stream'

    if settings.DEBUG:
        return FileResponse(evidencia.imagen.open('rb'), content_type=content_type)

    response = HttpResponse(content_type=content_type)
    response['X-Accel-Redirect'] = f'/protected-media/{evidencia.imagen.name}'
    return response


# ── AYUDA ──────────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def ayuda_sistema(request):
    return render(request, 'modulo_puntos/ayuda.html', {
        'rol_usuario': obtener_rol_usuario(request.user),
    })


# ── GESTIÓN DE USUARIOS ────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def crear_usuario_sistema(request):
    if not (usuario_es_superusuario(request.user) or usuario_es_admin_tic(request.user)):
        messages.error(request, 'No tienes permisos para crear usuarios del sistema.')
        return redirect('modulo_puntos:panel_control')

    pvd_pendiente = None
    pvd_pendiente_id = request.session.get('pvd_pendiente_id')
    if pvd_pendiente_id:
        pvd_pendiente = PuntoViveDigital.objects.filter(pk=pvd_pendiente_id).first()
        if not pvd_pendiente:
            request.session.pop('pvd_pendiente_id', None)
            request.session.pop('pvd_pendiente_nombre', None)

    # Admin TIC solo puede crear Admin PVD; si hay un PVD recién creado esperando
    # administrador, se fuerza el rol Admin PVD también para el Superusuario.
    solo_pvd = (not usuario_es_superusuario(request.user)) or pvd_pendiente is not None

    initial = {'pvd_asignado': pvd_pendiente.pk} if pvd_pendiente else None
    form = CrearUsuarioSistemaForm(request.POST or None, solo_pvd=solo_pvd, initial=initial)
    if request.method == 'POST':
        if form.is_valid():
            rol_elegido = form.cleaned_data['rol']
            # Doble verificación: si eligió admin_tic pero no es super → error
            if rol_elegido == 'admin_tic' and solo_pvd:
                messages.error(request, 'Solo el Superusuario puede crear Administradores TIC.')
                return redirect('modulo_puntos:crear_usuario_sistema')

            user = form.save()
            nombre_grupo = 'Administrador TIC' if rol_elegido == 'admin_tic' else 'Administrador PVD'
            grupo, _ = Group.objects.get_or_create(name=nombre_grupo)
            user.groups.add(grupo)

            pvd_asignado = form.cleaned_data.get('pvd_asignado')
            if rol_elegido == 'admin_pvd' and pvd_asignado:
                UserProfile.objects.update_or_create(
                    usuario=user,
                    defaults={'rol': 'admin_pvd', 'punto_asignado': pvd_asignado},
                )
                sincronizar_admin_a_cargo(user, pvd_nuevo=pvd_asignado)

            registrar_auditoria(request, 'CREATE', 'User', user.pk,
                                f'{nombre_grupo} creado: {user.username} — {user.get_full_name()}')

            mensaje = (
                f'{nombre_grupo} creado correctamente. '
                f'Usuario generado: "{user.username}"'
            )
            if pvd_asignado:
                mensaje += f'. Asignado al PVD "{pvd_asignado.nombre}".'
            messages.success(request, mensaje)

            if pvd_pendiente:
                request.session.pop('pvd_pendiente_id', None)
                request.session.pop('pvd_pendiente_nombre', None)
                if pvd_asignado and pvd_asignado.pk == pvd_pendiente.pk:
                    messages.success(request, f'El PVD "{pvd_pendiente.nombre}" ya cuenta con su administrador.')
                return redirect('modulo_puntos:lista_pvd')

            return redirect('modulo_puntos:accesos_temporales')
        messages.error(request, 'No se pudo crear el usuario. Revisa los datos ingresados.')

    return render(request, 'modulo_puntos/crear_usuario_sistema.html', {
        'form': form,
        'solo_pvd': solo_pvd,
        'pvd_pendiente': pvd_pendiente,
    })


@login_required(login_url='/login/')
def crear_admin_tic(request):
    return redirect('modulo_puntos:crear_usuario_sistema')


@login_required(login_url='/login/')
def crear_admin_pvd(request):
    return redirect('modulo_puntos:crear_usuario_sistema')


@login_required(login_url='/login/')
def lista_admins_pvd(request):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos para gestionar administradores PVD.')
        return redirect('modulo_puntos:panel_control')

    admins_pvd = (
        User.objects
        .filter(groups__name='Administrador PVD')
        .select_related('pvd_profile__punto_asignado')
        .distinct()
        .order_by('last_name', 'first_name', 'username')
    )

    return render(request, 'modulo_puntos/lista_admins_pvd.html', {
        'admins_pvd': admins_pvd,
    })


@login_required(login_url='/login/')
def editar_admin_pvd(request, user_id):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos para editar administradores PVD.')
        return redirect('modulo_puntos:panel_control')

    admin_pvd = get_object_or_404(User, pk=user_id, groups__name='Administrador PVD')
    profile, _ = UserProfile.objects.get_or_create(usuario=admin_pvd, defaults={'rol': 'admin_pvd'})

    form = EditarAdminPvdForm(
        request.POST or None,
        instance=admin_pvd,
        initial={'pvd_asignado': profile.punto_asignado_id},
    )
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            pvd_anterior = profile.punto_asignado
            profile.punto_asignado = form.cleaned_data['pvd_asignado']
            profile.save(update_fields=['punto_asignado'])
            sincronizar_admin_a_cargo(admin_pvd, pvd_anterior=pvd_anterior, pvd_nuevo=profile.punto_asignado)

            nueva_pwd = form.cleaned_data.get('password1')
            if nueva_pwd:
                admin_pvd.set_password(nueva_pwd)
                admin_pvd.save(update_fields=['password'])

            registrar_auditoria(request, 'UPDATE', 'User', admin_pvd.pk,
                                f'Administrador PVD editado: {admin_pvd.username}')
            messages.success(
                request,
                f'Administrador "{admin_pvd.get_full_name() or admin_pvd.username}" actualizado correctamente.'
            )
            return redirect('modulo_puntos:lista_admins_pvd')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/editar_admin_pvd.html', {
        'form': form,
        'admin_pvd': admin_pvd,
    })


# ── GESTIÓN DE ROLES Y PERMISOS ────────────────────────────────────────────────

@user_passes_test(lambda u: u.is_superuser)
def gestionar_roles(request):
    roles = Group.objects.all()
    permisos_disponibles = Permission.objects.filter(
        content_type__app_label='modulo_puntos_app'
    ).order_by('content_type__model')

    if request.method == 'POST':
        nombre_rol = request.POST.get('nombre_rol')
        permisos_id = request.POST.getlist('permisos')
        rol_id = request.POST.get('rol_id')

        if rol_id:
            rol = get_object_or_404(Group, id=rol_id)
            rol.name = nombre_rol
            rol.save()
        else:
            rol, _ = Group.objects.get_or_create(name=nombre_rol)

        rol.permissions.set(Permission.objects.filter(id__in=permisos_id))
        messages.success(request, f"Configuración de '{nombre_rol}' guardada.")
        return redirect('modulo_puntos:gestionar_roles')

    return render(request, 'modulo_puntos/gestionar_roles.html', {
        'roles': roles,
        'permisos_disponibles': permisos_disponibles
    })


@user_passes_test(lambda u: u.is_superuser)
def asignar_rol_usuario(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    grupos = Group.objects.all()

    if request.method == 'POST':
        grupo_id = request.POST.get('grupo_id')
        if grupo_id:
            grupo = get_object_or_404(Group, pk=grupo_id)
            usuario.groups.set([grupo])
            messages.success(request, f'Rol "{grupo.name}" asignado a {usuario.username}.')
        else:
            usuario.groups.clear()
            messages.success(request, f'Roles removidos de {usuario.username}.')
        return redirect('modulo_puntos:gestionar_roles')

    permisos_disponibles = Permission.objects.filter(
        content_type__app_label='modulo_puntos_app'
    ).order_by('content_type__model')

    return render(request, 'modulo_puntos/gestionar_roles.html', {
        'usuario_objetivo': usuario,
        'grupos': grupos,
        'roles': grupos,
        'permisos_disponibles': permisos_disponibles,
    })


@user_passes_test(lambda u: u.is_superuser)
def crear_grupo_rol(request):
    if request.method == 'POST':
        nombre_rol = request.POST.get('nombre_rol', '').strip()
        if nombre_rol:
            grupo, created = Group.objects.get_or_create(name=nombre_rol)
            if created:
                messages.success(request, f'Rol "{nombre_rol}" creado correctamente.')
            else:
                messages.warning(request, f'El rol "{nombre_rol}" ya existe.')
    return redirect('modulo_puntos:gestionar_roles')


# ── GESTIÓN DE PUNTOS VIVE DIGITAL ─────────────────────────────────────────────

@login_required(login_url='/login/')
def validar_nombre_pvd(request):
    nombre = request.GET.get('nombre', '').strip()
    exclude_id = request.GET.get('exclude_id', '').strip()
    if not nombre:
        return JsonResponse({'disponible': True})
    qs = PuntoViveDigital.objects.filter(nombre__iexact=nombre)
    if exclude_id:
        try:
            qs = qs.exclude(pk=int(exclude_id))
        except ValueError:
            pass
    return JsonResponse({'disponible': not qs.exists()})


@login_required(login_url='/login/')
def lista_pvd(request):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos para gestionar PVDs.')
        return redirect('modulo_puntos:panel_control')
    puede_eliminar_pvd = tiene_permiso(request.user, 'infraestructura.eliminar_pvd')

    # Conteos en una sola query con annotate
    pvds = (
        PuntoViveDigital.objects
        .order_by('nombre')
        .annotate(
            total_atenciones=Count('atencion', distinct=True),
            total_ciudadanos=Count(
                'ciudadano',
                filter=Q(ciudadano__estado='A'),
                distinct=True,
            ),
        )
    )

    # Admins por PVD en una sola query
    from .models import UserProfile
    profiles = UserProfile.objects.select_related('usuario', 'punto_asignado').filter(punto_asignado__isnull=False)
    admin_por_pvd = {}
    for p in profiles:
        admin_por_pvd.setdefault(p.punto_asignado_id, []).append(p.usuario)

    pvd_list = []
    for pvd in pvds:
        pvd.admins = admin_por_pvd.get(pvd.pk, [])
        pvd_list.append(pvd)

    return render(request, 'modulo_puntos/lista_pvd.html', {
        'pvds': pvd_list,
        'puede_eliminar_pvd': puede_eliminar_pvd,
    })


@login_required(login_url='/login/')
def crear_pvd(request):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos para crear PVDs.')
        return redirect('modulo_puntos:panel_control')

    initial = {'nombre': 'PVD '}
    form = PuntoViveDigitalForm(request.POST or None, initial=initial)
    if request.method == 'POST':
        if form.is_valid():
            pvd = form.save()
            registrar_auditoria(request, 'CREATE', 'PuntoViveDigital', pvd.pk, f'PVD creado: {pvd.nombre}')
            request.session['pvd_pendiente_id'] = pvd.pk
            request.session['pvd_pendiente_nombre'] = pvd.nombre
            messages.success(
                request,
                f'PVD "{pvd.nombre}" creado correctamente. Ahora crea el Administrador PVD que lo tendrá a cargo.'
            )
            return redirect('modulo_puntos:crear_usuario_sistema')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/form_pvd.html', {
        'form': form,
        'titulo': 'Nuevo Punto Vive Digital',
        'accion': 'crear',
    })


@login_required(login_url='/login/')
def editar_pvd(request, pvd_cdgo):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos para editar PVDs.')
        return redirect('modulo_puntos:panel_control')

    pvd = get_object_or_404(PuntoViveDigital, pk=pvd_cdgo)
    form = PuntoViveDigitalForm(request.POST or None, instance=pvd)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            cache.delete('pvds_disponibles_activos')
            registrar_auditoria(request, 'UPDATE', 'PuntoViveDigital', pvd.pk, f'PVD editado: {pvd.nombre}')
            messages.success(request, f'PVD "{pvd.nombre}" actualizado correctamente.')
            return redirect('modulo_puntos:lista_pvd')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/form_pvd.html', {
        'form': form,
        'titulo': f'Editar PVD: {pvd.nombre}',
        'accion': 'editar',
        'pvd': pvd,
    })


@login_required(login_url='/login/')
def activar_pvd(request, pvd_cdgo):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if request.method != 'POST':
        return redirect('modulo_puntos:lista_pvd')

    pvd = get_object_or_404(PuntoViveDigital, pk=pvd_cdgo)
    pvd.estado = 'I' if pvd.estado == 'A' else 'A'
    pvd.save()
    estado_str = 'activado' if pvd.estado == 'A' else 'desactivado'
    cache.delete('pvds_disponibles_activos')
    registrar_auditoria(request, 'UPDATE', 'PuntoViveDigital', pvd.pk, f'PVD {estado_str}: {pvd.nombre}')
    messages.success(request, f'PVD "{pvd.nombre}" {estado_str} correctamente.')
    return redirect('modulo_puntos:lista_pvd')


@login_required(login_url='/login/')
def eliminar_pvd(request, pvd_cdgo):
    messages.error(request, 'La eliminación de Puntos Vive Digital está deshabilitada. Usa la opción de activar/desactivar.')
    return redirect('modulo_puntos:lista_pvd')


# ── GESTIÓN DE SALAS ───────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def lista_salas(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'salas.ver'):
        messages.error(request, 'No tienes permiso para ver salas.')
        return redirect('modulo_puntos:panel_control')

    es_admin_tic = usuario_es_admin_tic(request.user) or request.user.is_superuser
    puede_eliminar = tiene_permiso(request.user, 'infraestructura.eliminar_sala')

    if es_admin_tic:
        pvds = PuntoViveDigital.objects.filter(estado='A').order_by('nombre')
        pvd_id = request.GET.get('pvd_id') or request.session.get('pvd_activo_id')
        pvd_seleccionado = PuntoViveDigital.objects.filter(pk=pvd_id).first() if pvd_id else None
        if pvd_seleccionado:
            salas = Sala.objects.filter(punto_vive_digital=pvd_seleccionado).order_by('nombre')
        else:
            salas = Sala.objects.all().select_related('punto_vive_digital').order_by('punto_vive_digital__nombre', 'nombre')
    else:
        pvds = []
        pvd_id = request.session.get('pvd_activo_id')
        pvd_seleccionado = PuntoViveDigital.objects.filter(pk=pvd_id).first() if pvd_id else None
        salas = Sala.objects.filter(punto_vive_digital_id=pvd_id).order_by('nombre') if pvd_id else Sala.objects.none()

    return render(request, 'modulo_puntos/lista_salas.html', {
        'salas': salas,
        'pvd': pvd_seleccionado,
        'pvd_seleccionado': pvd_seleccionado,
        'pvds': pvds,
        'es_admin_tic': es_admin_tic,
        'puede_eliminar': puede_eliminar,
    })


@login_required(login_url='/login/')
def crear_sala(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    # Solo superusuario o Admin TIC pueden crear salas
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'Solo el Superusuario o el Administrador TIC pueden crear nuevas salas.')
        return redirect('modulo_puntos:lista_salas')
    if not tiene_permiso(request.user, 'salas.crear'):
        messages.error(request, 'No tienes permiso para crear salas.')
        return redirect('modulo_puntos:lista_salas')

    pvd_id = request.session.get('pvd_activo_id')
    form = SalaForm(request.POST or None, initial={'punto_vive_digital': pvd_id} if pvd_id else None)
    if request.method == 'POST':
        if form.is_valid():
            sala = form.save()
            registrar_auditoria(request, 'CREATE', 'Sala', sala.pk, f'Sala creada: {sala.nombre}')
            messages.success(request, f'Sala "{sala.nombre}" creada correctamente.')
            return redirect('modulo_puntos:lista_salas')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/form_sala.html', {
        'form': form,
        'titulo': 'Crear Sala',
        'accion': 'crear',
    })


@login_required(login_url='/login/')
def editar_sala(request, sala_cdgo):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'salas.editar'):
        messages.error(request, 'No tienes permiso para editar salas.')
        return redirect('modulo_puntos:lista_salas')

    sala = get_object_or_404(Sala, pk=sala_cdgo)
    if not pvd_permitido(request, sala.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para editar salas de otro PVD.')
        return redirect('modulo_puntos:lista_salas')
    form = SalaForm(request.POST or None, instance=sala)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            registrar_auditoria(request, 'UPDATE', 'Sala', sala.pk, f'Sala editada: {sala.nombre}')
            messages.success(request, f'Sala "{sala.nombre}" actualizada correctamente.')
            return redirect('modulo_puntos:lista_salas')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/form_sala.html', {
        'form': form,
        'titulo': f'Editar Sala: {sala.nombre}',
        'accion': 'editar',
        'sala': sala,
    })


@login_required(login_url='/login/')
def activar_sala(request, sala_cdgo):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'salas.editar'):
        messages.error(request, 'No tienes permiso para cambiar el estado de salas.')
        return redirect('modulo_puntos:lista_salas')

    if request.method != 'POST':
        return redirect('modulo_puntos:lista_salas')
    sala = get_object_or_404(Sala, pk=sala_cdgo)
    if not pvd_permitido(request, sala.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para cambiar el estado de salas de otro PVD.')
        return redirect('modulo_puntos:lista_salas')
    sala.estado = 'I' if sala.estado == 'A' else 'A'
    sala.save()
    estado_str = 'activada' if sala.estado == 'A' else 'desactivada'
    registrar_auditoria(request, 'UPDATE', 'Sala', sala.pk, f'Sala {estado_str}: {sala.nombre}')
    messages.success(request, f'Sala "{sala.nombre}" {estado_str} correctamente.')
    return redirect('modulo_puntos:lista_salas')


@login_required(login_url='/login/')
def eliminar_sala(request, sala_cdgo):
    messages.error(request, 'La eliminación de salas está deshabilitada. Usa la opción de desactivar para retirarla.')
    return redirect('modulo_puntos:lista_salas')


# ── MÓDULO PERMISOS ────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def lista_permisos_roles(request):
    es_superusuario = request.user.is_superuser
    es_admin_tic = not es_superusuario and usuario_es_admin_tic(request.user)

    if not (es_superusuario or es_admin_tic):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('modulo_puntos:panel_control')

    roles = [('admin_tic', 'Administrador TIC'), ('admin_pvd', 'Administrador PVD')]
    usuarios_con_permisos = User.objects.filter(
        groups__name__in=['Administrador TIC', 'Administrador PVD']
    ).distinct().order_by('last_name', 'first_name', 'username')

    permisos = PermisoDefinicion.objects.filter(activo=True).order_by('categoria', 'nombre')

    if request.method == 'POST':
        todos_permisos = list(PermisoDefinicion.objects.filter(activo=True))
        for permiso in todos_permisos:
            for rol_codigo, _ in roles:
                checkbox_name = f'perm_{permiso.pk}_{rol_codigo}'
                tiene = checkbox_name in request.POST
                if tiene:
                    PermisoRol.objects.get_or_create(
                        rol=rol_codigo,
                        permiso=permiso,
                        defaults={'otorgado_por': request.user},
                    )
                else:
                    PermisoRol.objects.filter(rol=rol_codigo, permiso=permiso).delete()

            if es_superusuario:
                delegable = f'delegable_{permiso.pk}' in request.POST
                if delegable != permiso.delegable_por_ofitic:
                    permiso.delegable_por_ofitic = delegable
                    permiso.save(update_fields=['delegable_por_ofitic'])

        actor = 'superusuario' if es_superusuario else 'admin_tic'
        registrar_auditoria(
            request, 'UPDATE', 'PermisoRol', None,
            f'Matriz de permisos por rol actualizada por {actor}.'
        )
        messages.success(request, 'Permisos actualizados correctamente.')
        return redirect('modulo_puntos:lista_permisos_roles')

    asignaciones = set(PermisoRol.objects.values_list('permiso_id', 'rol'))

    roles_data = []
    for rol_codigo, rol_nombre in roles:
        categorias = {}
        for permiso in permisos:
            cat = permiso.categoria
            if cat not in categorias:
                categorias[cat] = []
            categorias[cat].append({
                'permiso': permiso,
                'marcado': (permiso.pk, rol_codigo) in asignaciones,
                'field_name': f'perm_{permiso.pk}_{rol_codigo}',
            })
        roles_data.append({'codigo': rol_codigo, 'nombre': rol_nombre, 'categorias': categorias})

    delegables_data = []
    if es_superusuario:
        categorias_del = {}
        for permiso in permisos:
            cat = permiso.categoria
            if cat not in categorias_del:
                categorias_del[cat] = []
            categorias_del[cat].append(permiso)
        delegables_data = categorias_del

    return render(request, 'modulo_puntos/permisos/lista_roles.html', {
        'roles_data': roles_data,
        'usuarios_con_permisos': usuarios_con_permisos,
        'es_superusuario': es_superusuario,
        'delegables_data': delegables_data,
    })



@login_required(login_url='/login/')
@user_passes_test(lambda u: u.is_superuser)
def editar_permiso(request, permiso_id):
    """Editar definición de permiso existente. Solo superusuario."""
    permiso = get_object_or_404(PermisoDefinicion, pk=permiso_id)
    form = PermisoDefinicionForm(request.POST or None, instance=permiso)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            registrar_auditoria(
                request, 'UPDATE', 'PermisoDefinicion', permiso.pk,
                f'Permiso editado: [{permiso.categoria}] {permiso.nombre} ({permiso.codigo})'
            )
            messages.success(request, f'Permiso "{permiso.nombre}" actualizado correctamente.')
            return redirect('modulo_puntos:lista_permisos_roles')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/permisos/form_permiso.html', {
        'form': form,
        'titulo': f'Editar Permiso: {permiso.nombre}',
        'accion': 'editar',
        'permiso': permiso,
    })


@login_required(login_url='/login/')
def permisos_usuario(request, user_id):
    """Overrides individuales por usuario. Superusuario: cualquier usuario. Admin TIC: solo admin_pvd."""
    es_superusuario = request.user.is_superuser
    es_admin_tic = not es_superusuario and usuario_es_admin_tic(request.user)

    if not (es_superusuario or es_admin_tic):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('modulo_puntos:panel_control')

    usuario_objetivo = get_object_or_404(User, pk=user_id)

    grupos = list(usuario_objetivo.groups.values_list('name', flat=True))
    if 'Administrador TIC' in grupos:
        rol_usuario = 'admin_tic'
    elif 'Administrador PVD' in grupos:
        rol_usuario = 'admin_pvd'
    else:
        rol_usuario = 'admin_pvd'

    if es_admin_tic and rol_usuario == 'admin_tic':
        messages.error(request, 'No tienes permisos para gestionar los permisos de un Administrador TIC.')
        return redirect('modulo_puntos:lista_permisos_roles')

    if es_admin_tic:
        permisos_pvd_ids = PermisoRol.objects.filter(
            rol='admin_pvd'
        ).values_list('permiso_id', flat=True)
        permisos = PermisoDefinicion.objects.filter(
            activo=True, pk__in=permisos_pvd_ids
        ).order_by('categoria', 'nombre')
    else:
        permisos = PermisoDefinicion.objects.filter(activo=True).order_by('categoria', 'nombre')

    permisos_rol = set(
        PermisoRol.objects.filter(rol=rol_usuario).values_list('permiso_id', flat=True)
    )
    overrides_existentes = {
        pu.permiso_id: pu for pu in PermisoUsuario.objects.filter(usuario=usuario_objetivo)
    }

    if request.method == 'POST':
        for permiso in permisos:
            checkbox_name = f'perm_{permiso.pk}'
            marcado = checkbox_name in request.POST
            override_existente = overrides_existentes.get(permiso.pk)
            hereda_del_rol = permiso.pk in permisos_rol

            if marcado == hereda_del_rol and override_existente:
                override_existente.delete()
            elif marcado != hereda_del_rol:
                PermisoUsuario.objects.update_or_create(
                    usuario=usuario_objetivo,
                    permiso=permiso,
                    defaults={'concedido': marcado, 'otorgado_por': request.user},
                )
            elif marcado == hereda_del_rol and not override_existente:
                pass

        registrar_auditoria(
            request, 'UPDATE', 'PermisoUsuario', usuario_objetivo.pk,
            f'Permisos individuales actualizados para usuario: {usuario_objetivo.username}'
        )
        messages.success(request, f'Permisos de {usuario_objetivo.username} actualizados.')
        return redirect('modulo_puntos:permisos_usuario', user_id=user_id)

    categorias = {}
    for permiso in permisos:
        cat = permiso.categoria
        if cat not in categorias:
            categorias[cat] = []
        override = overrides_existentes.get(permiso.pk)
        if override is not None:
            activo = override.concedido
            es_override = True
        else:
            activo = permiso.pk in permisos_rol
            es_override = False
        categorias[cat].append({
            'permiso': permiso,
            'activo': activo,
            'es_override': es_override,
            'hereda_rol': permiso.pk in permisos_rol,
        })

    return render(request, 'modulo_puntos/permisos/usuario_permisos.html', {
        'usuario_objetivo': usuario_objetivo,
        'categorias': categorias,
        'rol_usuario': rol_usuario,
        'es_superusuario': es_superusuario,
    })


# ── HABILITACIÓN DE SALAS ──────────────────────────────────────────────────────

def _actualizar_estados_habilitaciones(base_qs):
    """Actualiza automáticamente el estado de habilitaciones según la hora actual."""
    ahora = datetime.now()
    hoy = ahora.date()
    hora_actual = ahora.time()

    base = base_qs.exclude(estado__in=('X', 'F'))

    # Fechas pasadas → Finalizado
    base.filter(fecha__lt=hoy).update(estado='F')

    # Hoy, hora en rango → En curso
    base.filter(
        fecha=hoy,
        hora_inicio__lte=hora_actual,
        hora_fin__gte=hora_actual,
    ).exclude(estado='E').update(estado='E')

    # Hoy, hora superada → Finalizado
    base.filter(
        fecha=hoy,
        hora_fin__lt=hora_actual,
    ).update(estado='F')


@login_required(login_url='/login/')
def lista_habilitaciones(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'habilitaciones.ver'):
        messages.error(request, 'No tienes permiso para ver habilitaciones de sala.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    fecha_filtro = request.GET.get('fecha', '')
    sala_filtro = request.GET.get('sala_id', '')
    estado_filtro = request.GET.get('estado', '')

    if pvd_id:
        salas_pvd = Sala.objects.filter(punto_vive_digital_id=pvd_id)
        qs = HabilitacionSala.objects.filter(sala__punto_vive_digital_id=pvd_id)
        pvd = PuntoViveDigital.objects.filter(pk=pvd_id).first()
    else:
        salas_pvd = Sala.objects.all().select_related('punto_vive_digital')
        qs = HabilitacionSala.objects.all()
        pvd = None

    qs = qs.select_related('sala', 'sala__punto_vive_digital', 'registrado_por')

    if fecha_filtro:
        try:
            qs = qs.filter(fecha=fecha_filtro)
        except Exception:
            pass
    if sala_filtro:
        qs = qs.filter(sala_id=sala_filtro)
    if estado_filtro:
        qs = qs.filter(estado=estado_filtro)

    _actualizar_estados_habilitaciones(qs)
    qs = qs.order_by('fecha', 'hora_inicio')

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'modulo_puntos/habilitaciones/lista_habilitaciones.html', {
        'habilitaciones': page_obj,
        'page_obj': page_obj,
        'salas_pvd': salas_pvd,
        'pvd': pvd,
        'fecha_filtro': fecha_filtro,
        'sala_filtro': sala_filtro,
        'estado_filtro': estado_filtro,
        'estados': HabilitacionSala.ESTADO_CHOICES,
    })


@login_required(login_url='/login/')
def crear_habilitacion(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'habilitaciones.crear'):
        messages.error(request, 'No tienes permiso para crear habilitaciones.')
        return redirect('modulo_puntos:lista_habilitaciones')

    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(request, 'Debes ingresar a un Punto Vive Digital antes de crear una habilitación de sala.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    sala_inicial = request.GET.get('sala_id')
    fecha_inicial = request.GET.get('fecha')
    atencion_id = request.GET.get('atencion_id') or request.POST.get('atencion_id')
    atencion_origen = Atencion.objects.filter(pk=atencion_id).first() if atencion_id else None

    initial = {}
    if sala_inicial:
        initial['sala'] = sala_inicial
    if fecha_inicial:
        initial['fecha'] = fecha_inicial
    if atencion_origen and atencion_origen.ciudadano_id:
        initial['solicitante'] = atencion_origen.ciudadano_id

    form = HabilitacionSalaForm(
        request.POST or None,
        pvd_id=pvd_id,
        initial=initial,
    )

    if request.method == 'POST':
        if form.is_valid():
            hab = form.save(commit=False)
            hab.registrado_por = request.user
            hab.save()
            form.save_m2m()
            registrar_auditoria(
                request, 'CREATE', 'HabilitacionSala', hab.pk,
                f'Habilitación creada: {hab.sala.nombre} – {hab.fecha}'
            )
            messages.success(request, f'Habilitación para "{hab.sala.nombre}" el {hab.fecha} registrada correctamente.')
            if atencion_id:
                return redirect('modulo_puntos:detalle_atencion', atencion_id=atencion_id)
            return redirect('modulo_puntos:lista_habilitaciones')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/habilitaciones/form_habilitacion.html', {
        'form': form,
        'titulo': 'Nueva Habilitación de Sala',
        'accion': 'crear',
        'atencion_id': atencion_id,
        'atencion_origen': atencion_origen,
    })


@login_required(login_url='/login/')
def editar_habilitacion(request, hab_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'habilitaciones.editar'):
        messages.error(request, 'No tienes permiso para editar habilitaciones.')
        return redirect('modulo_puntos:lista_habilitaciones')

    hab = get_object_or_404(HabilitacionSala.objects.select_related('sala'), pk=hab_id)
    if not pvd_permitido(request, hab.sala.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para editar habilitaciones de otro PVD.')
        return redirect('modulo_puntos:lista_habilitaciones')

    pvd_id = request.session.get('pvd_activo_id')
    form = HabilitacionSalaForm(request.POST or None, instance=hab, pvd_id=pvd_id)

    if request.method == 'POST':
        if form.is_valid():
            hab = form.save()
            registrar_auditoria(
                request, 'UPDATE', 'HabilitacionSala', hab.pk,
                f'Habilitación editada: {hab.sala.nombre} – {hab.fecha}'
            )
            messages.success(request, f'Habilitación actualizada correctamente.')
            return redirect('modulo_puntos:lista_habilitaciones')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/habilitaciones/form_habilitacion.html', {
        'form': form,
        'titulo': f'Editar Habilitación – {hab.sala.nombre}',
        'accion': 'editar',
        'hab': hab,
    })


@login_required(login_url='/login/')
def cancelar_habilitacion(request, hab_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'habilitaciones.cancelar'):
        messages.error(request, 'No tienes permiso para cancelar habilitaciones.')
        return redirect('modulo_puntos:lista_habilitaciones')

    try:
        hab = HabilitacionSala.objects.select_related('sala').get(pk=hab_id)
    except HabilitacionSala.DoesNotExist:
        messages.warning(request, 'La habilitación ya no existe.')
        return redirect('modulo_puntos:lista_habilitaciones')
    if not pvd_permitido(request, hab.sala.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para cancelar habilitaciones de otro PVD.')
        return redirect('modulo_puntos:lista_habilitaciones')

    if request.method != 'POST':
        return redirect('modulo_puntos:lista_habilitaciones')

    if hab.estado in ('F', 'X'):
        messages.warning(request, 'Esta habilitación ya está finalizada o cancelada.')
        return redirect('modulo_puntos:lista_habilitaciones')

    hab.estado = 'X'
    hab.save()
    registrar_auditoria(
        request, 'UPDATE', 'HabilitacionSala', hab.pk,
        f'Habilitación cancelada: {hab.sala.nombre} – {hab.fecha}'
    )
    messages.success(request, f'Habilitación de "{hab.sala.nombre}" cancelada.')
    return redirect('modulo_puntos:lista_habilitaciones')


@login_required(login_url='/login/')
def eliminar_habilitacion(request, hab_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'habilitaciones.eliminar'):
        messages.error(request, 'No tienes permiso para eliminar habilitaciones.')
        return redirect('modulo_puntos:lista_habilitaciones')

    if request.method != 'POST':
        return redirect('modulo_puntos:lista_habilitaciones')

    try:
        hab = HabilitacionSala.objects.select_related('sala').get(pk=hab_id)
    except HabilitacionSala.DoesNotExist:
        messages.warning(request, 'La habilitación ya no existe.')
        return redirect('modulo_puntos:lista_habilitaciones')
    if not pvd_permitido(request, hab.sala.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para eliminar habilitaciones de otro PVD.')
        return redirect('modulo_puntos:lista_habilitaciones')
    nombre_sala = hab.sala.nombre
    fecha = hab.fecha
    pk = hab.pk
    hab.delete()
    registrar_auditoria(
        request, 'DELETE', 'HabilitacionSala', pk,
        f'Habilitación eliminada: {nombre_sala} – {fecha}'
    )
    messages.success(request, f'Habilitación de "{nombre_sala}" eliminada.')
    return redirect('modulo_puntos:lista_habilitaciones')


@login_required(login_url='/login/')
def agenda_sala(request, sala_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')

    sala = get_object_or_404(Sala, pk=sala_id)
    if not pvd_permitido(request, sala.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para ver la agenda de salas de otro PVD.')
        return redirect('modulo_puntos:lista_salas')

    from datetime import date as dt_date, timedelta
    fecha_str = request.GET.get('fecha', '')
    try:
        fecha_base = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else dt_date.today()
    except ValueError:
        fecha_base = dt_date.today()

    fecha_inicio_semana = fecha_base - timedelta(days=fecha_base.weekday())
    dias_semana = [fecha_inicio_semana + timedelta(days=i) for i in range(7)]

    habilitaciones_semana = HabilitacionSala.objects.filter(
        sala=sala,
        fecha__range=[dias_semana[0], dias_semana[-1]],
    ).order_by('fecha', 'hora_inicio')

    _actualizar_estados_habilitaciones(habilitaciones_semana)

    agenda = {d: [] for d in dias_semana}
    for hab in habilitaciones_semana:
        if hab.fecha in agenda:
            agenda[hab.fecha].append(hab)

    semana_anterior = (fecha_inicio_semana - timedelta(days=7)).strftime('%Y-%m-%d')
    semana_siguiente = (fecha_inicio_semana + timedelta(days=7)).strftime('%Y-%m-%d')

    return render(request, 'modulo_puntos/habilitaciones/agenda_sala.html', {
        'sala': sala,
        'dias_semana': dias_semana,
        'agenda': agenda,
        'semana_anterior': semana_anterior,
        'semana_siguiente': semana_siguiente,
        'fecha_base': fecha_base,
    })


# ==============================================================================
# CURSOS / TALLERES
# ==============================================================================

@login_required(login_url='/login/')
def lista_cursos(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.ver'):
        messages.error(request, 'No tienes permiso para ver cursos.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    es_admin_tic = usuario_es_admin_tic(request.user)

    if es_admin_tic:
        qs = Curso.objects.select_related('punto_vive_digital', 'registrado_por').all()
    elif pvd_id:
        qs = Curso.objects.filter(punto_vive_digital_id=pvd_id).select_related('punto_vive_digital', 'registrado_por')
    else:
        qs = Curso.objects.none()
    qs = qs.annotate(total_inscritos=Count('inscripciones', filter=Q(inscripciones__estado__in=['I', 'C'])))

    estado_filtro = request.GET.get('estado', '')
    fecha_desde_filtro = request.GET.get('fecha_desde', '').strip()
    fecha_hasta_filtro = request.GET.get('fecha_hasta', '').strip()
    q_filtro = request.GET.get('q', '').strip()
    if estado_filtro:
        qs = qs.filter(estado=estado_filtro)
    if fecha_desde_filtro:
        qs = qs.filter(fecha_inicio__gte=fecha_desde_filtro)
    if fecha_hasta_filtro:
        qs = qs.filter(fecha_inicio__lte=fecha_hasta_filtro)
    if q_filtro:
        qs = qs.filter(
            Q(nombre__icontains=q_filtro) |
            Q(poblacion_objetivo__icontains=q_filtro) |
            Q(descripcion__icontains=q_filtro)
        )

    paginator = Paginator(qs.order_by('-fecha_inicio'), 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'modulo_puntos/cursos/lista_cursos.html', {
        'cursos': page_obj,
        'page_obj': page_obj,
        'estado_filtro': estado_filtro,
        'fecha_desde_filtro': fecha_desde_filtro,
        'fecha_hasta_filtro': fecha_hasta_filtro,
        'q_filtro': q_filtro,
        'estados': Curso.ESTADO_CHOICES,
        'es_admin_tic': es_admin_tic,
        'total_cursos': qs.count(),
    })


@login_required(login_url='/login/')
def crear_curso(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.crear'):
        messages.error(request, 'No tienes permiso para crear cursos.')
        return redirect('modulo_puntos:lista_cursos')

    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.error(request, 'Debes seleccionar un Punto Vive Digital primero.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    form = CursoForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            curso = form.save(commit=False)
            if pvd_id:
                curso.punto_vive_digital_id = pvd_id
            curso.registrado_por = request.user
            curso.save()
            registrar_auditoria(request, 'CREATE', 'Curso', curso.pk, f'Curso creado: {curso.nombre}')
            messages.success(request, f'Curso "{curso.nombre}" creado correctamente.')
            return redirect('modulo_puntos:detalle_curso', curso_id=curso.pk)
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/cursos/form_curso.html', {
        'form': form,
        'titulo': 'Nuevo Curso / Taller',
        'accion': 'crear',
    })


@login_required(login_url='/login/')
def editar_curso(request, curso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.editar'):
        messages.error(request, 'No tienes permiso para editar cursos.')
        return redirect('modulo_puntos:lista_cursos')

    curso = get_object_or_404(Curso, pk=curso_id)
    if not pvd_permitido(request, curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para editar cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    form = CursoForm(request.POST or None, instance=curso)

    if request.method == 'POST':
        if form.is_valid():
            curso = form.save()
            registrar_auditoria(request, 'UPDATE', 'Curso', curso.pk, f'Curso editado: {curso.nombre}')
            messages.success(request, 'Curso actualizado correctamente.')
            return redirect('modulo_puntos:detalle_curso', curso_id=curso.pk)
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/cursos/form_curso.html', {
        'form': form,
        'titulo': f'Editar: {curso.nombre}',
        'accion': 'editar',
        'curso': curso,
    })


@login_required(login_url='/login/')
def detalle_curso(request, curso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.ver'):
        messages.error(request, 'No tienes permiso para ver cursos.')
        return redirect('modulo_puntos:panel_control')

    curso = get_object_or_404(Curso.objects.select_related('punto_vive_digital', 'registrado_por'), pk=curso_id)
    if not pvd_permitido(request, curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para ver cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    sesiones = curso.sesiones.order_by('numero_sesion')
    inscripciones = curso.inscripciones.select_related('ciudadano').order_by('fecha_inscripcion')

    return render(request, 'modulo_puntos/cursos/detalle_curso.html', {
        'curso': curso,
        'sesiones': sesiones,
        'inscripciones': inscripciones,
    })


@login_required(login_url='/login/')
def crear_sesion_curso(request, curso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.sesiones'):
        messages.error(request, 'No tienes permiso para gestionar sesiones de cursos.')
        return redirect('modulo_puntos:lista_cursos')

    curso = get_object_or_404(Curso, pk=curso_id)
    if not pvd_permitido(request, curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para gestionar sesiones de cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    ultimo_num = curso.sesiones.aggregate(max_n=Max('numero_sesion'))['max_n'] or 0
    form = SesionCursoForm(request.POST or None, initial={'numero_sesion': ultimo_num + 1})

    if request.method == 'POST':
        if form.is_valid():
            sesion = form.save(commit=False)
            sesion.curso = curso
            sesion.save()
            registrar_auditoria(request, 'CREATE', 'SesionCurso', sesion.pk,
                                f'Sesión {sesion.numero_sesion} creada para curso {curso.nombre}')
            messages.success(request, f'Sesión {sesion.numero_sesion} registrada.')
            return redirect('modulo_puntos:detalle_curso', curso_id=curso.pk)
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/cursos/form_sesion.html', {
        'form': form,
        'curso': curso,
        'titulo': f'Nueva Sesión — {curso.nombre}',
    })


@login_required(login_url='/login/')
def inscribir_ciudadano(request, curso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.inscribir'):
        messages.error(request, 'No tienes permiso para inscribir ciudadanos en cursos.')
        return redirect('modulo_puntos:lista_cursos')

    curso = get_object_or_404(Curso, pk=curso_id)
    if not pvd_permitido(request, curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para inscribir ciudadanos en cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    inscritos_ids = InscripcionCurso.objects.filter(curso=curso).values_list('ciudadano_id', flat=True)
    ciudadanos_disponibles = Ciudadano.objects.filter(estado='A').exclude(pk__in=inscritos_ids)

    form = InscripcionCursoForm(request.POST or None)
    form.fields['ciudadano'].queryset = ciudadanos_disponibles

    if request.method == 'POST':
        if form.is_valid():
            inscripcion = form.save(commit=False)
            inscripcion.curso = curso
            inscripcion.registrado_por = request.user
            inscripcion.save()
            registrar_auditoria(request, 'CREATE', 'InscripcionCurso', inscripcion.pk,
                                f'{inscripcion.ciudadano} inscrito en {curso.nombre}')
            messages.success(request, f'{inscripcion.ciudadano} inscrito correctamente.')
            return redirect('modulo_puntos:detalle_curso', curso_id=curso.pk)
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/cursos/form_inscripcion.html', {
        'form': form,
        'curso': curso,
        'titulo': f'Inscribir Ciudadano — {curso.nombre}',
    })


@login_required(login_url='/login/')
def marcar_asistencia(request, sesion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.asistencia'):
        messages.error(request, 'No tienes permiso para registrar asistencia.')
        return redirect('modulo_puntos:lista_cursos')

    sesion = get_object_or_404(SesionCurso.objects.select_related('curso'), pk=sesion_id)
    if not pvd_permitido(request, sesion.curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para registrar asistencia en cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    inscritos = InscripcionCurso.objects.filter(
        curso=sesion.curso, estado__in=['I', 'C']
    ).select_related('ciudadano')

    asistencias_existentes = {
        a.ciudadano_id: a for a in AsistenciaSesion.objects.filter(sesion=sesion)
    }

    if request.method == 'POST':
        try:
            presentes = {int(v) for v in request.POST.getlist('asistio') if str(v).strip().isdigit()}
        except (ValueError, TypeError):
            presentes = set()
        for insc in inscritos:
            cid = insc.ciudadano_id
            asistio = cid in presentes
            if cid in asistencias_existentes:
                a = asistencias_existentes[cid]
                if a.asistio != asistio:
                    a.asistio = asistio
                    a.save()
            else:
                AsistenciaSesion.objects.create(sesion=sesion, ciudadano=insc.ciudadano, asistio=asistio)
        registrar_auditoria(request, 'UPDATE', 'SesionCurso', sesion.pk,
                            f'Asistencia registrada para sesión {sesion.numero_sesion} de {sesion.curso.nombre}')
        messages.success(request, 'Asistencia guardada correctamente.')
        return redirect('modulo_puntos:detalle_curso', curso_id=sesion.curso.pk)

    lista = []
    for insc in inscritos:
        lista.append({
            'ciudadano': insc.ciudadano,
            'asistio': asistencias_existentes.get(insc.ciudadano_id, None) and asistencias_existentes[insc.ciudadano_id].asistio,
        })

    return render(request, 'modulo_puntos/cursos/asistencia_sesion.html', {
        'sesion': sesion,
        'lista': lista,
    })


@login_required(login_url='/login/')
def cambiar_estado_curso(request, curso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.editar'):
        messages.error(request, 'No tienes permiso para editar cursos.')
        return redirect('modulo_puntos:lista_cursos')
    if request.method != 'POST':
        return redirect('modulo_puntos:lista_cursos')

    curso = get_object_or_404(Curso, pk=curso_id)
    if not pvd_permitido(request, curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para cambiar el estado de cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    nuevo_estado = (request.POST.get('nuevo_estado') or request.POST.get('estado', '')).strip()
    estados_validos = {'PL', 'AC', 'FI', 'CA'}
    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado no válido.')
    else:
        curso.estado = nuevo_estado
        curso.save(update_fields=['estado'])
        registrar_auditoria(request, 'UPDATE', 'Curso', curso.pk,
                            f'Estado cambiado a {curso.get_estado_display()}')
        messages.success(request, f'Curso marcado como "{curso.get_estado_display()}".')
    return redirect('modulo_puntos:detalle_curso', curso_id=curso.pk)


@login_required(login_url='/login/')
def eliminar_curso(request, curso_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.editar'):
        messages.error(request, 'No tienes permiso para eliminar cursos.')
        return redirect('modulo_puntos:lista_cursos')
    curso = get_object_or_404(Curso, pk=curso_id)
    if not pvd_permitido(request, curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para eliminar cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    if request.method == 'POST':
        nombre = curso.nombre
        registrar_auditoria(request, 'DELETE', 'Curso', curso.pk, f'Curso eliminado: {nombre}')
        curso.delete()
        messages.success(request, f'Curso "{nombre}" eliminado.')
        return redirect('modulo_puntos:lista_cursos')
    return redirect('modulo_puntos:detalle_curso', curso_id=curso.pk)


@login_required(login_url='/login/')
def eliminar_sesion_curso(request, sesion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'cursos.sesiones'):
        messages.error(request, 'No tienes permiso para gestionar sesiones.')
        return redirect('modulo_puntos:lista_cursos')
    sesion = get_object_or_404(SesionCurso.objects.select_related('curso'), pk=sesion_id)
    curso_id = sesion.curso_id
    if not pvd_permitido(request, sesion.curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para eliminar sesiones de cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    if request.method == 'POST':
        registrar_auditoria(request, 'DELETE', 'SesionCurso', sesion.pk,
                            f'Sesión {sesion.numero_sesion} eliminada del curso #{curso_id}')
        sesion.delete()
        messages.success(request, f'Sesión {sesion.numero_sesion} eliminada.')
    return redirect('modulo_puntos:detalle_curso', curso_id=curso_id)


@login_required(login_url='/login/')
def eliminar_inscripcion(request, inscripcion_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('modulo_puntos:panel_control')
    inscripcion = get_object_or_404(InscripcionCurso.objects.select_related('curso'), pk=inscripcion_id)
    curso_id = inscripcion.curso_id
    if not pvd_permitido(request, inscripcion.curso.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para eliminar inscripciones de cursos de otro PVD.')
        return redirect('modulo_puntos:lista_cursos')
    if request.method == 'POST':
        nombre = inscripcion.ciudadano.get_nombre_completo() if inscripcion.ciudadano else '—'
        registrar_auditoria(request, 'DELETE', 'InscripcionCurso', inscripcion.pk,
                            f'Inscripción eliminada: {nombre} del curso #{curso_id}')
        inscripcion.delete()
        messages.success(request, f'Inscripción de {nombre} eliminada.')
    return redirect('modulo_puntos:detalle_curso', curso_id=curso_id)


# ==============================================================================
# MANTENIMIENTO DE EQUIPOS
# ==============================================================================

@login_required(login_url='/login/')
def lista_mantenimientos(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'mantenimiento.ver'):
        messages.error(request, 'No tienes permiso para ver mantenimientos.')
        return redirect('modulo_puntos:panel_control')

    pvd_id = request.session.get('pvd_activo_id')
    es_admin_tic = usuario_es_admin_tic(request.user)

    if es_admin_tic:
        qs = MantenimientoEquipo.objects.select_related('punto_vive_digital', 'realizado_por').all()
    elif pvd_id:
        qs = MantenimientoEquipo.objects.filter(punto_vive_digital_id=pvd_id).select_related('punto_vive_digital', 'realizado_por')
    else:
        qs = MantenimientoEquipo.objects.none()

    tipo_filtro = request.GET.get('tipo', '')
    fecha_desde_filtro = request.GET.get('fecha_desde', '').strip()
    fecha_hasta_filtro = request.GET.get('fecha_hasta', '').strip()
    q_filtro = request.GET.get('q', '').strip()
    if tipo_filtro:
        qs = qs.filter(tipo=tipo_filtro)
    if fecha_desde_filtro:
        qs = qs.filter(fecha__gte=fecha_desde_filtro)
    if fecha_hasta_filtro:
        qs = qs.filter(fecha__lte=fecha_hasta_filtro)
    if q_filtro:
        qs = qs.filter(
            Q(equipos_intervenidos__icontains=q_filtro) |
            Q(descripcion__icontains=q_filtro) |
            Q(realizado_por__first_name__icontains=q_filtro) |
            Q(realizado_por__last_name__icontains=q_filtro)
        )

    paginator = Paginator(qs.order_by('-fecha'), 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'modulo_puntos/mantenimientos/lista_mantenimientos.html', {
        'mantenimientos': page_obj,
        'page_obj': page_obj,
        'tipo_filtro': tipo_filtro,
        'fecha_desde_filtro': fecha_desde_filtro,
        'fecha_hasta_filtro': fecha_hasta_filtro,
        'q_filtro': q_filtro,
        'tipos': MantenimientoEquipo.TIPO_CHOICES,
        'es_admin_tic': es_admin_tic,
        'total_mantenimientos': qs.count(),
    })


@login_required(login_url='/login/')
def crear_mantenimiento(request):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    # Solo el Admin PVD y el Superusuario pueden registrar mantenimientos
    _es_admin_pvd = (
        not request.user.is_superuser
        and not usuario_es_admin_tic(request.user)
        and request.user.groups.filter(name='Administrador PVD').exists()
    )
    if not _es_admin_pvd and not request.user.is_superuser:
        messages.error(request, 'Solo el Administrador PVD o el Superusuario pueden registrar mantenimientos de equipos.')
        return redirect('modulo_puntos:lista_mantenimientos')
    if not tiene_permiso(request.user, 'mantenimiento.crear'):
        messages.error(request, 'No tienes permiso para registrar mantenimientos.')
        return redirect('modulo_puntos:lista_mantenimientos')

    pvd_id = request.session.get('pvd_activo_id')
    if not pvd_id:
        messages.warning(request, 'Debes ingresar a un Punto Vive Digital antes de registrar un mantenimiento.')
        return redirect('modulo_puntos:seleccionar_pvd_view')

    form = MantenimientoEquipoForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            mant = form.save(commit=False)
            if pvd_id:
                mant.punto_vive_digital_id = pvd_id
            mant.realizado_por = request.user
            mant.save()
            registrar_auditoria(request, 'CREATE', 'MantenimientoEquipo', mant.pk,
                                f'Mantenimiento {mant.get_tipo_display()} registrado: {mant.fecha}')
            messages.success(request, 'Mantenimiento registrado correctamente.')
            return redirect('modulo_puntos:lista_mantenimientos')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/mantenimientos/form_mantenimiento.html', {
        'form': form,
        'titulo': 'Registrar Mantenimiento',
        'accion': 'crear',
    })


@login_required(login_url='/login/')
def editar_mantenimiento(request, mant_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'mantenimiento.editar'):
        messages.error(request, 'No tienes permiso para editar mantenimientos.')
        return redirect('modulo_puntos:lista_mantenimientos')

    mant = get_object_or_404(MantenimientoEquipo, pk=mant_id)
    if not pvd_permitido(request, mant.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para editar mantenimientos de otro PVD.')
        return redirect('modulo_puntos:lista_mantenimientos')

    form = MantenimientoEquipoForm(request.POST or None, instance=mant)

    if request.method == 'POST':
        if form.is_valid():
            mant = form.save()
            registrar_auditoria(request, 'UPDATE', 'MantenimientoEquipo', mant.pk,
                                f'Mantenimiento editado: {mant.fecha}')
            messages.success(request, 'Mantenimiento actualizado correctamente.')
            return redirect('modulo_puntos:lista_mantenimientos')
        messages.error(request, 'Revisa los datos del formulario.')

    return render(request, 'modulo_puntos/mantenimientos/form_mantenimiento.html', {
        'form': form,
        'titulo': f'Editar Mantenimiento — {mant.fecha}',
        'accion': 'editar',
        'mant': mant,
    })


@login_required(login_url='/login/')
def eliminar_mantenimiento(request, mant_id):
    if not usuario_puede_usar_modulos_pvd(request.user):
        messages.error(request, 'No tienes permisos para acceder a este módulo.')
        return redirect('modulo_puntos:panel_control')
    if not tiene_permiso(request.user, 'mantenimiento.editar'):
        messages.error(request, 'No tienes permiso para eliminar mantenimientos.')
        return redirect('modulo_puntos:lista_mantenimientos')

    mant = get_object_or_404(MantenimientoEquipo, pk=mant_id)
    if not pvd_permitido(request, mant.punto_vive_digital_id):
        messages.error(request, 'No tienes permiso para eliminar mantenimientos de otro PVD.')
        return redirect('modulo_puntos:lista_mantenimientos')

    if request.method == 'POST':
        registrar_auditoria(request, 'DELETE', 'MantenimientoEquipo', mant.pk,
                            f'Mantenimiento eliminado: {mant.get_tipo_display()} — {mant.fecha}')
        mant.delete()
        messages.success(request, 'Registro de mantenimiento eliminado.')

    return redirect('modulo_puntos:lista_mantenimientos')


# ── ACCESOS TEMPORALES ─────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def accesos_temporales(request):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('modulo_puntos:panel_control')

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        pvd_id = request.POST.get('pvd_id') or None
        campo = request.POST.get('campo', 'pvd_temporal')

        try:
            usuario_objetivo = User.objects.get(pk=user_id, groups__name='Administrador PVD')
        except User.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
            return redirect('modulo_puntos:accesos_temporales')

        profile, _ = UserProfile.objects.get_or_create(
            usuario=usuario_objetivo,
            defaults={'rol': 'admin_pvd'}
        )

        nombre_usuario = usuario_objetivo.get_full_name() or usuario_objetivo.username

        if campo == 'punto_asignado':
            pvd_anterior = profile.punto_asignado
            if pvd_id:
                try:
                    pvd = PuntoViveDigital.objects.get(pk=pvd_id)
                    profile.punto_asignado = pvd
                    profile.save(update_fields=['punto_asignado'])
                    sincronizar_admin_a_cargo(usuario_objetivo, pvd_anterior=pvd_anterior, pvd_nuevo=pvd)
                    messages.success(request, f'PVD asignado a {nombre_usuario}: {pvd.nombre}')
                except PuntoViveDigital.DoesNotExist:
                    messages.error(request, 'PVD no encontrado.')
            else:
                profile.punto_asignado = None
                profile.save(update_fields=['punto_asignado'])
                sincronizar_admin_a_cargo(usuario_objetivo, pvd_anterior=pvd_anterior, pvd_nuevo=None)
                messages.success(request, f'PVD permanente de {nombre_usuario} eliminado.')
        else:
            if pvd_id:
                try:
                    pvd = PuntoViveDigital.objects.get(pk=pvd_id, estado='A')
                    profile.pvd_temporal = pvd
                    profile.save(update_fields=['pvd_temporal'])
                    messages.success(request, f'Acceso temporal a "{pvd.nombre}" otorgado a {nombre_usuario}.')
                except PuntoViveDigital.DoesNotExist:
                    messages.error(request, 'PVD no encontrado o inactivo.')
            else:
                nombre_pvd = profile.pvd_temporal.nombre if profile.pvd_temporal else '—'
                profile.pvd_temporal = None
                profile.save(update_fields=['pvd_temporal'])
                messages.success(request, f'Acceso temporal de {nombre_usuario} a "{nombre_pvd}" revocado.')

        registrar_auditoria(
            request, 'UPDATE', 'UserProfile', profile.pk,
            f'{request.user.username} modificó acceso PVD de {nombre_usuario}'
        )
        return redirect('modulo_puntos:accesos_temporales')

    admins_pvd = (
        User.objects
        .filter(groups__name='Administrador PVD')
        .select_related('pvd_profile__punto_asignado', 'pvd_profile__pvd_temporal')
        .distinct()
        .order_by('last_name', 'first_name', 'username')
    )
    pvds_activos = PuntoViveDigital.objects.filter(estado='A').order_by('nombre')

    return render(request, 'modulo_puntos/accesos_temporales.html', {
        'admins_pvd': admins_pvd,
        'pvds_activos': pvds_activos,
    })


# ── LOG DE AUDITORÍA ───────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def log_auditoria(request):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'Solo Administradores TIC y Superusuarios pueden ver el log de auditoría.')
        return redirect('modulo_puntos:panel_control')

    q = request.GET.get('q', '').strip()
    accion_filter = request.GET.get('accion', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    qs = AuditoriaAccion.objects.all()
    if q:
        qs = qs.filter(
            Q(usuario__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(modelo_afectado__icontains=q)
        )
    if accion_filter:
        qs = qs.filter(accion=accion_filter)
    if fecha_desde:
        qs = qs.filter(fecha_accion__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_accion__date__lte=fecha_hasta)

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'modulo_puntos/log_auditoria.html', {
        'registros': page_obj,
        'page_obj': page_obj,
        'q': q,
        'accion_filter': accion_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_choices': AuditoriaAccion.TIPO_ACCION,
    })


@login_required(login_url='/login/')
def exportar_auditoria(request):
    if not usuario_es_admin_tic(request.user):
        messages.error(request, 'Solo Administradores TIC y Superusuarios pueden exportar el log de auditoría.')
        return redirect('modulo_puntos:panel_control')

    q = request.GET.get('q', '').strip()
    accion_filter = request.GET.get('accion', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    qs = AuditoriaAccion.objects.all()
    if q:
        qs = qs.filter(
            Q(usuario__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(modelo_afectado__icontains=q)
        )
    if accion_filter:
        qs = qs.filter(accion=accion_filter)
    if fecha_desde:
        qs = qs.filter(fecha_accion__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_accion__date__lte=fecha_hasta)

    headers = ['Fecha / Hora', 'Usuario', 'Acción', 'Modelo', 'ID Objeto', 'Descripción', 'IP']
    filas = []
    for reg in qs:
        filas.append([
            reg.fecha_accion.strftime('%d/%m/%Y %H:%M') if reg.fecha_accion else '',
            reg.usuario or '',
            reg.get_accion_display(),
            reg.modelo_afectado or '',
            str(reg.objeto_id) if reg.objeto_id else '',
            reg.descripcion or '',
            reg.direccion_ip or '',
        ])

    registrar_auditoria(request, 'EXPORT', 'AuditoriaAccion', None, f'Exportación log auditoría ({len(filas)} registros)')
    wb = _crear_hoja('Log Auditoría', headers, filas)
    return _xlsx_response('Log_Auditoria_PVD', wb)


